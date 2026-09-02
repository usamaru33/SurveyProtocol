#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""判定を速くするためのローカル閲覧アプリを生成する。

**このツールは判定を一切行わない。** キーボード1打で「人が下した判定」を記録するだけ。
自動で Include / Exclude を書き込む機能は意図的に持たせていない。理由:

  κ は校正セット(calibration=Y)で **著者 × 各評価者** のペアについて算出する
  (`score_screening.py` の「ペアごとの Cohen's κ」)。著者側の判定が正規表現の
  出力だと、κ は「評価者間の一致」ではなく「人 vs 正規表現の一致」を測ることになり、
  報告値が無意味になる。rule.md §プロトコル改訂(2026-07-16) が
  「意味的判断は人手ダブルスクリーニング」と定めているのもこのため。

  したがって本ツールが提供するのは **読む速度** であって **判断の代行** ではない。
  キーワードは「並べ替え」と「ハイライト」にのみ使う(下記 --rules)。

入力:  screening/sheet_<id>.csv (雛形。既に xlsx に記入があればそれも読む)
出力:  screening/review_<id>.html  … 単体で開けるHTML(依存なし・オフライン動作)

使い方:
    python -X utf8 scripts/make_review_app.py                # 著者用
    python -X utf8 scripts/make_review_app.py --id kataoka   # 評価者用
    python -X utf8 scripts/make_review_app.py --rules my_rules.json

生成した HTML をブラウザで開き、キーボードで判定する。判定は逐次 localStorage に
保存され、`E` キーで decisions CSV を書き出す。CSV を xlsx へ反映するのは
`scripts/apply_review_decisions.py`。

途中経過は `S` キーで JSON へ書き出し、`L` キーで読み戻せる(既定の保存先は
`screening/json/progress_<id>.json`)。localStorage はブラウザの閲覧データ削除で
消えるため、JSON がその唯一のバックアップになる。JSON は
`apply_review_decisions.py --input ...json` でそのまま xlsx へ反映できる。
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import sys
from pathlib import Path


def _sha1(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()

sys.path.insert(0, str(Path(__file__).resolve().parent))
csv.field_size_limit(10 ** 9)

# 統制語彙は xlsx 生成側が正。ここで定義し直すと必ず乖離する。
from make_screening_xlsx import DECISIONS, EXCLUDE_REASONS, REASON_OTHER  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SCREENING = ROOT / "screening"

# 複数該当時の優先順位 (Rev.24)。正は docs/reference/reviewer_briefing.md §3.3。
#   上位ほど「下位の基準を評価する前提そのものが崩れている」順に並ぶ。
#   キー 1-9 の並びは語彙の表示順(EXCLUDE_REASONS)で凍結されているため一致しない。
#   凡例に順位を併記するのはこのズレで取り違えるのを防ぐため。
EXCLUDE_PRIORITY = [
    "重複(同一研究の別報告)",
    "スコープ外(主題が無関係)",
    "S: 原著論文でない",
    "S: ユーザー実験が無い",
    "P: 対象者が不適合",
    "I: HMD-VR でない",
    "I: スケール操作/多感覚刺激が無い",
    "O: スケール知覚の測定が無い",
    "その他",
]

# 語彙が変わったのに優先順位を直し忘れると、凡例が黙って嘘をつく。
_vocab = {v for v, _ in EXCLUDE_REASONS}
if set(EXCLUDE_PRIORITY) != _vocab:
    raise SystemExit(
        "EXCLUDE_PRIORITY が統制語彙と一致しません。\n"
        f"  優先順位にのみ存在: {sorted(set(EXCLUDE_PRIORITY) - _vocab)}\n"
        f"  語彙にのみ存在:     {sorted(_vocab - set(EXCLUDE_PRIORITY))}\n"
        "docs/reference/reviewer_briefing.md §3.3 を確認して両方を揃えてください。")

# ハイライト語。
#   group1-3 は実行済み検索クエリの3概念群 (docs/protocol/search_strings.md Rev.6)。
#   cue は PICOS の判断材料になりやすい語。**両方向(残す手がかり/落とす手がかり)を
#   対称に入れてある。**片側だけ光らせると誘導になるため。
HIGHLIGHT = {
    "g1": ["virtual reality", "VR", "HMD", "head-mounted display", "head mounted display",
           "virtual environment", "immersive virtual", "immersive"],
    "g2": ["avatar", "body", "embodiment", "self-avatar", "body ownership", "virtual body"],
    "g3": ["size", "scale", "scaling", "height", "distance", "eye height", "proportion"],
    "cue": ["participants", "participant", "subjects", "user study", "we conducted",
            "experiment", "within-subjects", "between-subjects", "questionnaire",
            "desktop", "monitor", "CAVE", "augmented reality", "AR", "mixed reality",
            "patients", "children", "elderly", "clinical", "rehabilitation", "surgeon",
            "we present", "we propose", "prototype", "demonstration",
            "survey", "review", "state of the art"],
}


def fingerprint(records: list[dict]) -> str:
    """レコード集合の指紋。「件数:ID列のSHA-1先頭7桁」。

    進捗JSON がこの HTML と同じ判定対象に対するものかを読み込み時に照合する。
    件数だけだと入れ替え(同数で別集合)を見逃し、ハッシュだけだと人が見て何が
    違うのか分からないので両方を出す。
    """
    ids = [str(r.get("record_id") or "") for r in records]
    return "{}:{}".format(len(ids), _sha1("\n".join(ids))[:7])


def load_records(sheet_id: str) -> list[dict]:
    """雛形 CSV を読み、同名 xlsx に既存の記入があれば取り込む。"""
    csv_path = SCREENING / f"sheet_{sheet_id}.csv"
    if not csv_path.exists():
        sys.exit(f"[ERROR] {csv_path} が無い。先に make_screening_sheets.py を実行すること。")
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        rows = [r for r in csv.DictReader(f) if r.get("record_id")]

    # 校正セットの正は assignment.csv。sheet_*.csv は生成時点のスナップショットなので、
    # Rev.22(校正セット 15%→20%)のようにあとから割当が変わるとズレる。
    # ★ の表示を誤ると「κ に効く文献」を見落とすので、必ず assignment 側で上書きする。
    apath = SCREENING / "assignment.csv"
    if apath.exists():
        with apath.open(encoding="utf-8-sig", newline="") as f:
            cal = {r["record_id"]: r.get("calibration", "N")
                   for r in csv.DictReader(f) if r.get("record_id")}
        drift = 0
        for r in rows:
            want = cal.get(r["record_id"], r.get("calibration", "N"))
            if want != r.get("calibration"):
                r["calibration"] = want
                drift += 1
        if drift:
            print(f"[INFO] assignment.csv により校正セット表示を {drift} 件補正した"
                  f"（sheet_{sheet_id}.csv が割当より古い）。")

    xlsx_path = SCREENING / f"sheet_{sheet_id}.xlsx"
    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("[WARN] openpyxl が無いため xlsx の既存記入は取り込まない。")
        else:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb["判定"]
            it = ws.iter_rows(values_only=True)
            header = list(next(it))
            idx = {name: header.index(name) for name in ("ID", "判定 ★", "除外理由 ★", "メモ")
                   if name in header}
            existing = {}
            for row in it:
                rid = row[idx["ID"]] if idx.get("ID") is not None else None
                if not rid:
                    continue
                existing[str(rid)] = {
                    "decision": row[idx["判定 ★"]] or "",
                    "reason": row[idx["除外理由 ★"]] or "",
                    "note": row[idx["メモ"]] or "",
                }
            wb.close()
            n = 0
            for r in rows:
                got = existing.get(r["record_id"])
                if got and str(got["decision"]).strip():
                    r["decision"], r["reason"], r["note"] = (
                        str(got["decision"]).strip(), str(got["reason"] or "").strip(),
                        str(got["note"] or "").strip())
                    n += 1
            if n:
                print(f"[INFO] xlsx から既存の判定 {n} 件を取り込んだ。")
    return rows


def build_payload(rows: list[dict], translations: dict | None = None) -> list[dict]:
    keep = ("record_id", "title", "abstract", "venue", "year", "rank", "has_abstract",
            "source", "calibration", "abstract_source", "kw_groups", "doi",
            "decision", "reason", "note")
    tr = translations or {}
    out = []
    for r in rows:
        rec = {k: (r.get(k) or "") for k in keep}
        got = tr.get(r["record_id"])
        # 訳文は原文が一致するものだけ載せる。原文が差し替わったのに古い訳が
        # 残っていると、読んでいる訳と判定対象がズレる。
        if got and got.get("src_sha1") == _sha1((r.get("abstract") or "").strip()):
            rec["ja"] = got.get("ja", "")
            rec["ja_engine"] = got.get("engine", "")
        out.append(rec)
    return out


def render(sheet_id: str, records: list[dict], rules: list[dict]) -> str:
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    reasons = json.dumps([{"value": v, "desc": d} for v, d in EXCLUDE_REASONS],
                         ensure_ascii=False)
    return TEMPLATE.replace("__SHEET_ID__", html.escape(sheet_id)) \
                   .replace("__RECORDS__", payload) \
                   .replace("__REASONS__", reasons) \
                   .replace("__PRIORITY__", json.dumps(EXCLUDE_PRIORITY, ensure_ascii=False)) \
                   .replace("__DECISIONS__", json.dumps(DECISIONS, ensure_ascii=False)) \
                   .replace("__REASON_OTHER__", json.dumps(REASON_OTHER, ensure_ascii=False)) \
                   .replace("__HIGHLIGHT__", json.dumps(HIGHLIGHT, ensure_ascii=False)) \
                   .replace("__RULES__", json.dumps(rules, ensure_ascii=False)) \
                   .replace("__FINGERPRINT__", fingerprint(records))


TEMPLATE = r"""<!doctype html>
<html lang="ja"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Phase 3b 判定 — __SHEET_ID__</title>
<style>
:root{
  --bg:#ffffff; --fg:#1a1a1a; --muted:#6b7280; --line:#e5e7eb; --soft:#f6f8fa;
  --accent:#1f4e79; --inc:#0f7b3f; --exc:#b02a37; --uns:#b26a00;
  --g1:#dbeafe; --g2:#dcfce7; --g3:#fef3c7; --cue:#ede9fe; --cal:#fdf1e7;
}
@media (prefers-color-scheme:dark){:root:not([data-theme=light]){
  --bg:#111418; --fg:#e6e8eb; --muted:#9aa4b2; --line:#2a3038; --soft:#171b21;
  --accent:#7fb0e0; --inc:#4ade80; --exc:#f87171; --uns:#fbbf24;
  --g1:#1e3a5f; --g2:#14532d; --g3:#5c4708; --cue:#3b2f6b; --cal:#4a3520;
}}
*{box-sizing:border-box}
html,body{margin:0;padding:0;background:var(--bg);color:var(--fg);
  font-family:system-ui,"Segoe UI","Hiragino Kaku Gothic ProN","Yu Gothic UI",sans-serif;}
#bar{position:sticky;top:0;z-index:10;background:var(--bg);border-bottom:1px solid var(--line);
  padding:8px 16px;display:flex;gap:14px;align-items:center;flex-wrap:wrap;font-size:13px}
#prog{flex:1;min-width:160px;height:8px;background:var(--soft);border-radius:99px;overflow:hidden}
#progfill{height:100%;background:var(--accent);width:0%}
.pill{padding:2px 9px;border-radius:99px;background:var(--soft);white-space:nowrap}
.pill b{font-variant-numeric:tabular-nums}
main{max-width:920px;margin:0 auto;padding:22px 20px 120px}
.meta{display:flex;gap:10px;flex-wrap:wrap;align-items:center;font-size:12.5px;color:var(--muted);margin-bottom:10px}
.tag{padding:2px 8px;border-radius:4px;background:var(--soft);white-space:nowrap}
.tag.cal{background:var(--cal);color:var(--fg);font-weight:700}
.tag.noabs{background:var(--exc);color:#fff}
h1{font-size:21px;line-height:1.5;margin:6px 0 12px;font-weight:700}
.abs{font-size:15px;line-height:1.95;white-space:pre-wrap;
  border-left:3px solid var(--line);padding:2px 0 2px 16px;max-width:74ch}
.abs.empty{color:var(--muted);font-style:italic;border-left-color:var(--exc)}
.panelab{font-size:11.5px;letter-spacing:.04em;color:var(--muted);text-transform:uppercase;
  margin:18px 0 5px;font-weight:700}
.panelab .src{text-transform:none;letter-spacing:0;font-weight:400}
.abs.ja{border-left-color:var(--accent);background:var(--soft);border-radius:0 6px 6px 0;
  padding:10px 16px;line-height:1.9}
mark{padding:0 2px;border-radius:3px;background:var(--soft);color:inherit}
mark.g1{background:var(--g1)} mark.g2{background:var(--g2)}
mark.g3{background:var(--g3)} mark.cue{background:var(--cue)}
#state{margin:22px 0 8px;font-size:15px;font-weight:700}
#state .inc{color:var(--inc)} #state .exc{color:var(--exc)} #state .uns{color:var(--uns)}
#note{width:100%;max-width:74ch;font:inherit;font-size:14px;padding:8px 10px;
  border:1px solid var(--line);border-radius:6px;background:var(--bg);color:var(--fg)}
#keys{position:fixed;bottom:0;left:0;right:0;background:var(--bg);
  border-top:1px solid var(--line);padding:8px 16px;font-size:12px;color:var(--muted);
  display:flex;gap:8px;flex-wrap:wrap;justify-content:center}
kbd{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;background:var(--soft);
  border:1px solid var(--line);border-radius:4px;padding:1px 5px;font-size:11px;color:var(--fg)}
/* 常時表示のキー凡例。判定中に「どの数字がどの理由か」を毎回ヘルプで開かずに済ませる。 */
#legend{position:fixed;right:0;top:0;bottom:0;width:270px;overflow:auto;z-index:9;
  background:var(--soft);border-left:1px solid var(--line);
  padding:14px 14px 76px;font-size:12.5px;display:none}
body.legend-on #legend{display:block}
body.legend-on{padding-right:270px}
#legend h3{margin:0 0 6px;font-size:11px;letter-spacing:.05em;text-transform:uppercase;
  color:var(--muted);font-weight:700}
#legend h3.sec{margin-top:16px}
#legend .row{display:flex;gap:7px;align-items:baseline;padding:5px 0;
  border-bottom:1px solid var(--line)}
#legend .row:last-child{border-bottom:0}
#legend .lb{flex:1;line-height:1.5}
#legend .pr{color:var(--muted);font-size:10.5px;white-space:nowrap}
#legend .foot{margin-top:14px;color:var(--muted);line-height:1.65;font-size:11.5px}
/* 凡例を出しているときは下部バーの理由チップを畳む(同じ情報の二重表示を避ける)。 */
body.legend-on .rchip{display:none}
@media (max-width:1080px){
  body.legend-on{padding-right:0}
  body.legend-on #legend{display:none}
  body.legend-on .rchip{display:inline}
}
#help{position:fixed;inset:0;background:rgba(0,0,0,.55);display:none;z-index:50;
  align-items:center;justify-content:center;padding:20px}
#help.on{display:flex}
#helpbox{background:var(--bg);border-radius:10px;padding:22px 26px;max-width:640px;
  max-height:82vh;overflow:auto;border:1px solid var(--line)}
#helpbox table{border-collapse:collapse;width:100%;font-size:13.5px}
#helpbox td{padding:4px 8px;border-bottom:1px solid var(--line);vertical-align:top}
button{font:inherit;font-size:13px;padding:5px 11px;border:1px solid var(--line);
  border-radius:6px;background:var(--soft);color:var(--fg);cursor:pointer}
select{font:inherit;font-size:13px;padding:4px 6px;border:1px solid var(--line);
  border-radius:6px;background:var(--bg);color:var(--fg)}
.warn{background:var(--cal);border-radius:8px;padding:10px 14px;font-size:13px;margin-bottom:14px}
/* 進捗JSON の読み込みは判定を丸ごと置き換える。実行前に必ず内訳を見せる。 */
#dlg{position:fixed;inset:0;background:rgba(0,0,0,.55);display:none;z-index:60;
  align-items:center;justify-content:center;padding:20px}
#dlg.on{display:flex}
#dlgbox{background:var(--bg);border:1px solid var(--line);border-radius:10px;
  padding:20px 24px;max-width:600px;max-height:82vh;overflow:auto;
  font-size:13.5px;line-height:1.75}
#dlgbox h2{margin:0 0 10px;font-size:16px}
#dlgbox .dmsg{color:var(--muted);margin:8px 0;word-break:break-all}
#dlgbox .dtab{border-collapse:collapse;margin:12px 0;font-variant-numeric:tabular-nums}
#dlgbox .dtab td{padding:4px 26px 4px 0;border-bottom:1px solid var(--line)}
#dlgbox .dtab td:last-child{text-align:right;font-weight:700;padding-right:0}
#dlgbox .dwarn{background:var(--cal);border-radius:8px;padding:10px 13px;margin:12px 0}
#dlgbox .btns{display:flex;gap:8px;justify-content:flex-end;margin-top:18px;flex-wrap:wrap}
#dlgbox button.primary{background:var(--accent);color:#fff;border-color:var(--accent)}
#toast{position:fixed;left:50%;bottom:52px;transform:translateX(-50%);z-index:70;
  max-width:80vw;background:var(--fg);color:var(--bg);padding:9px 16px;border-radius:8px;
  font-size:13px;line-height:1.6;box-shadow:0 4px 16px rgba(0,0,0,.25);display:none}
#toast.on{display:block}
#toast.bad{background:var(--exc);color:#fff}
#jsonstat{cursor:default}
</style></head><body>

<div id="bar">
  <span class="pill">#<b id="pos">0</b>/<b id="total">0</b></span>
  <div id="prog"><div id="progfill"></div></div>
  <span class="pill">判定済 <b id="done">0</b></span>
  <span class="pill" style="color:var(--inc)">Inc <b id="ninc">0</b></span>
  <span class="pill" style="color:var(--exc)">Exc <b id="nexc">0</b></span>
  <span class="pill" style="color:var(--uns)">Uns <b id="nuns">0</b></span>
  <select id="filter">
    <option value="all">すべて</option>
    <option value="todo">未判定のみ</option>
    <option value="cal">校正セットのみ</option>
    <option value="noabs">要旨なしのみ</option>
    <option value="flag">ルール該当のみ</option>
  </select>
  <button id="jamode" hidden>訳 対訳</button>
  <span class="pill" id="jsonstat">JSON 未保存</span>
  <button id="exp">CSV書き出し (E)</button>
  <button id="sav">JSON保存 (S)</button>
  <button id="ld">JSON読込 (L)</button>
  <button id="lgd">凡例 (R)</button>
  <button id="hlp">? ヘルプ</button>
</div>

<main>
  <div id="calwarn" class="warn" hidden>
    <b>★ 校正セット</b> — この文献は3名全員が判定し、<b>Cohen's κ の算出に使われます。</b>
    普段どおりの基準で、ただし丁寧に読んでください。
  </div>
  <div class="meta" id="meta"></div>
  <h1 id="title"></h1>
  <div class="panelab" id="lab-en" hidden>原文（判定はこちらに対して行う）</div>
  <div class="abs" id="abs"></div>
  <div class="panelab" id="lab-ja" hidden>機械翻訳（参考） <span class="src" id="ja-eng"></span></div>
  <div class="abs ja" id="absja" hidden></div>
  <div id="state">未判定</div>
  <input id="note" placeholder="メモ（任意。除外理由が「その他」のときは必須）">
</main>

<div id="keys"></div>
<aside id="legend"></aside>
<div id="help"><div id="helpbox">
  <h2 style="margin-top:0">キー操作</h2>
  <table id="helptable"></table>
  <p style="font-size:13px;color:var(--muted);margin-bottom:0">
    判定は入力のたびにブラウザに保存されます（localStorage）。作業を終えたら
    <kbd>E</kbd> で CSV を書き出し、<code>scripts/apply_review_decisions.py</code> で
    xlsx に反映してください。<br><br>
    <b>途中保存は <kbd>S</kbd>。</b> 判定を JSON ファイル
    （<code>progress___SHEET_ID__.json</code>）へ書き出します。保存先フォルダは
    初回に一度だけ選びます（<code>screening/json</code> を選んでください）。2回目以降は
    同じフォルダへ上書きします。<kbd>L</kbd> でそのファイルを読み戻し、
    <kbd>Shift</kbd>+<kbd>L</kbd> でファイルを選んで読み込みます。<br>
    localStorage は<b>ブラウザの閲覧データを消すと一緒に消えます</b>。JSON がその唯一の
    バックアップなので、区切りごとに <kbd>S</kbd> を押してください。<br><br>
    <b>このツールは判定を自動化しません。</b> キーワードルールはハイライトと並べ替えだけに
    使われ、Include / Exclude を書き込むことはありません。
  </p>
</div></div>

<div id="dlg"><div id="dlgbox"></div></div>
<div id="toast"></div>
<!-- File System Access API が無いブラウザ用の読み込み口 -->
<input type="file" id="fileinput" accept=".json,application/json" hidden>

<script>
const RECORDS=__RECORDS__, REASONS=__REASONS__, DECISIONS=__DECISIONS__;
const PRIORITY=__PRIORITY__;
const REASON_OTHER=__REASON_OTHER__, HIGHLIGHT=__HIGHLIGHT__, RULES=__RULES__;
const SHEET="__SHEET_ID__", KEY="phase3b:"+SHEET;
// 進捗JSON。FINGERPRINT は判定対象の集合が同じかを読み込み時に照合するためのもの。
const FINGERPRINT="__FINGERPRINT__";
const PROGRESS_FORMAT="phase3b-progress", PROGRESS_VERSION=1;
const PROGRESS_FILE="progress_"+SHEET+".json";

let store={};
try{store=JSON.parse(localStorage.getItem(KEY)||"{}")}catch(e){store={}}
// 生成時に xlsx から取り込んだ判定は、localStorage が空のときだけ初期値にする。
RECORDS.forEach(r=>{ if(!store[r.record_id] && r.decision)
  store[r.record_id]={decision:r.decision,reason:r.reason||"",note:r.note||""}; });

function matchRules(r){
  const hay=((r.title||"")+" "+(r.abstract||"")+" "+(r.venue||"")).toLowerCase();
  return RULES.filter(ru=>(ru.any||[]).some(w=>hay.includes(String(w).toLowerCase())))
              .map(ru=>ru.label||"rule");
}
RECORDS.forEach(r=>{ r._flags=matchRules(r); });

let order=RECORDS.map((_,i)=>i), view=[], cur=0;
function rebuild(keepId){
  const f=document.getElementById("filter").value;
  view=order.filter(i=>{const r=RECORDS[i],s=store[r.record_id];
    if(f==="todo")  return !(s&&s.decision);
    if(f==="cal")   return r.calibration==="Y";
    if(f==="noabs") return r.has_abstract==="N";
    if(f==="flag")  return r._flags.length>0;
    return true;});
  if(!view.length){view=order.slice()}
  if(keepId!==undefined){const k=view.findIndex(i=>RECORDS[i].record_id===keepId); if(k>=0)cur=k;}
  cur=Math.max(0,Math.min(cur,view.length-1));
  draw();
}

function esc(s){return (s||"").replace(/[&<>]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;"}[c]))}
let hlOn=true;
// 0=原文のみ / 1=対訳 / 2=訳のみ。既定は対訳。
// 「訳のみ」を既定にしないのは、判定を原文に対して下すため(README §4.5)。
let jaMode=1;
const JA_MODE_LABEL=["原文のみ","対訳","訳のみ"];
function highlight(text){
  let out=esc(text);
  if(!hlOn) return out;
  for(const cls of ["g1","g2","g3","cue"]){
    const terms=[...HIGHLIGHT[cls]].sort((a,b)=>b.length-a.length);
    for(const t of terms){
      const pat=t.replace(/[.*+?^${}()|[\]\\]/g,"\\$&").replace(/[\s-]/g,"[\\s-]");
      out=out.replace(new RegExp("(?<![\\w>])("+pat+")(?![\\w])","gi"),
                      m=>'<mark class="'+cls+'">'+m+'</mark>');
    }
  }
  return out;
}

function draw(){
  const r=RECORDS[view[cur]], s=store[r.record_id]||{};
  document.getElementById("pos").textContent=cur+1;
  document.getElementById("total").textContent=view.length;
  const tags=[];
  if(r.calibration==="Y") tags.push('<span class="tag cal">★ 校正セット</span>');
  if(r.has_abstract==="N") tags.push('<span class="tag noabs">要旨なし</span>');
  tags.push('<span class="tag">'+esc(r.venue)+'</span>');
  tags.push('<span class="tag">'+esc(r.year)+'</span>');
  if(r.rank) tags.push('<span class="tag">'+esc(r.rank)+'</span>');
  tags.push('<span class="tag">概念群 '+esc(r.kw_groups)+'</span>');
  tags.push('<span class="tag">'+esc(r.source)+'</span>');
  if(r.abstract_source==="enriched") tags.push('<span class="tag">要旨=補完</span>');
  if(r.doi) tags.push('<span class="tag">'+esc(r.doi)+'</span>');
  r._flags.forEach(f=>tags.push('<span class="tag" style="background:var(--cue)">▶ '+esc(f)+'</span>'));
  document.getElementById("meta").innerHTML=tags.join("");
  document.getElementById("calwarn").hidden=(r.calibration!=="Y");
  document.getElementById("title").innerHTML=highlight(r.title);
  const a=document.getElementById("abs");
  if(r.abstract){a.className="abs";a.innerHTML=highlight(r.abstract);}
  else{a.className="abs empty";a.textContent="（要旨なし — タイトルだけで判断できなければ Unsure に）";}
  // 対訳。訳文だけの表示(mode=2)でも原文の見出しは残さない代わりに、
  // 訳文側に「参考」と明示し続ける。判定は原文に対して行うため。
  const hasJa=!!r.ja, showEn=(jaMode!==2), showJa=hasJa&&(jaMode!==0);
  document.getElementById("lab-en").hidden=!(hasJa&&showEn);
  document.getElementById("abs").hidden=!showEn;
  document.getElementById("lab-ja").hidden=!showJa;
  document.getElementById("absja").hidden=!showJa;
  document.getElementById("ja-eng").textContent=hasJa&&r.ja_engine?("／"+r.ja_engine):"";
  if(showJa) document.getElementById("absja").textContent=r.ja;
  const st=document.getElementById("state");
  if(!s.decision) st.innerHTML='<span style="color:var(--muted)">未判定</span>';
  else if(s.decision==="Include") st.innerHTML='<span class="inc">✓ Include</span>';
  else if(s.decision==="Unsure")  st.innerHTML='<span class="uns">? Unsure</span>';
  else st.innerHTML='<span class="exc">✕ Exclude</span> <span style="font-weight:400">— '+esc(s.reason)+'</span>';
  document.getElementById("note").value=s.note||"";
  stats(); window.scrollTo(0,0);
}

// 上部バーと進捗JSON の counts は同じ数え方でなければならない(ズレると
// 「ファイルには 201 件と書いてあるのに画面は 200 件」で原因追跡が始まる)。
function tally(src){
  const st=src||store; let done=0,inc=0,exc=0,uns=0;
  for(const r of RECORDS){const s=st[r.record_id]; if(!s||!s.decision)continue;
    done++; if(s.decision==="Include")inc++; else if(s.decision==="Exclude")exc++; else uns++;}
  return {done:done,include:inc,exclude:exc,unsure:uns,total:RECORDS.length};
}
function stats(){
  const t=tally();
  document.getElementById("done").textContent=t.done;
  document.getElementById("ninc").textContent=t.include;
  document.getElementById("nexc").textContent=t.exclude;
  document.getElementById("nuns").textContent=t.unsure;
  document.getElementById("progfill").style.width=(t.done/RECORDS.length*100).toFixed(1)+"%";
  drawSaved(t);
}

let undo=[];
function set(dec,reason){
  const r=RECORDS[view[cur]];
  undo.push({id:r.record_id,prev:JSON.parse(JSON.stringify(store[r.record_id]||{}))});
  if(undo.length>200)undo.shift();
  const s=store[r.record_id]||{};
  store[r.record_id]={decision:dec,reason:dec==="Exclude"?(reason||""):"",note:s.note||""};
  save(); draw();
  if(!(dec==="Exclude"&&reason===REASON_OTHER)) next();
  else document.getElementById("note").focus();
}
function save(){localStorage.setItem(KEY,JSON.stringify(store))}
function next(){if(cur<view.length-1){cur++;draw()}else{draw()}}
function prev(){if(cur>0){cur--;draw()}}

document.getElementById("note").addEventListener("input",e=>{
  const r=RECORDS[view[cur]];
  const s=store[r.record_id]||{decision:"",reason:""};
  s.note=e.target.value; store[r.record_id]=s; save();
});

function exportCSV(){
  const head=["record_id","decision","reason","note"];
  const lines=[head.join(",")];
  for(const r of RECORDS){
    const s=store[r.record_id]; if(!s||!s.decision) continue;
    lines.push([r.record_id,s.decision,s.reason||"",s.note||""]
      .map(v=>'"'+String(v).replace(/"/g,'""')+'"').join(","));
  }
  download("\ufeff"+lines.join("\r\n"),"decisions_"+SHEET+".csv","text/csv;charset=utf-8");
}
function download(text,name,type){
  const a=document.createElement("a");
  a.href=URL.createObjectURL(new Blob([text],{type:type}));
  a.download=name; a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),2000);
}

// ---- 途中経過の JSON 書き出し / 読み込み -------------------------------------
// localStorage はブラウザの閲覧データ削除・別PC・別ブラウザで失われる。1,052件を
// 読み直す羽目になるので、同じ内容をファイルへ退避できるようにする。
// 保存先フォルダ(screening/json/)は初回に一度だけ選ぶ。File System Access API の
// ハンドルを IndexedDB に覚えるので、2回目以降はダイアログなしで上書きできる。
// IndexedDB が使えない環境ではセッション中だけ覚え、API 自体が無ければ通常の
// ダウンロードへ落とす。どの経路でも保存自体は必ずできるようにしてある。
const FS_OK=(typeof window.showDirectoryPicker==="function");
let dirHandle=null;

function idbOpen(){return new Promise((res,rej)=>{
  let rq; try{rq=indexedDB.open("phase3b",1)}catch(e){return rej(e)}
  rq.onupgradeneeded=()=>rq.result.createObjectStore("kv");
  rq.onsuccess=()=>res(rq.result); rq.onerror=()=>rej(rq.error);});}
async function idbGet(k){try{const db=await idbOpen();
  return await new Promise((res,rej)=>{
    const rq=db.transaction("kv","readonly").objectStore("kv").get(k);
    rq.onsuccess=()=>res(rq.result||null); rq.onerror=()=>rej(rq.error);});
  }catch(e){return null}}
async function idbPut(k,v){try{const db=await idbOpen();
  await new Promise((res,rej)=>{const tx=db.transaction("kv","readwrite");
    tx.objectStore("kv").put(v,k); tx.oncomplete=()=>res(); tx.onerror=()=>rej(tx.error);});
  }catch(e){}}

async function perm(h){
  if(!h) return false;
  if(!h.queryPermission) return true;
  const o={mode:"readwrite"};
  try{ if(await h.queryPermission(o)==="granted") return true;
       return (await h.requestPermission(o))==="granted"; }catch(e){return false}
}
// ask=false なら「既に許可済みのフォルダ」だけを返す(勝手にダイアログを出さない)。
async function getDir(ask){
  if(!FS_OK) return null;
  if(!dirHandle) dirHandle=await idbGet("dir:"+SHEET);
  if(dirHandle && await perm(dirHandle)) return dirHandle;
  dirHandle=null;
  if(!ask) return null;
  const h=await window.showDirectoryPicker({id:"phase3b-json",mode:"readwrite"});
  dirHandle=h; await idbPut("dir:"+SHEET,h); return h;
}

function buildProgress(){
  // 判定した順ではなくシートの並びで書く。同じ状態なら同じファイルになり、
  // 2つの進捗JSON を diff で比べられる。
  const decisions={};
  for(const r of RECORDS){const s=store[r.record_id];
    if(!s||!s.decision) continue;
    decisions[r.record_id]={decision:s.decision,reason:s.reason||"",note:s.note||""};}
  return {format:PROGRESS_FORMAT,version:PROGRESS_VERSION,sheet:SHEET,
          saved_at:new Date().toISOString(),fingerprint:FINGERPRINT,
          counts:tally(),decisions:decisions};
}

function noteSaved(p,where){
  localStorage.setItem("json:"+SHEET,p.saved_at);
  localStorage.setItem("json:"+SHEET+":n",String(p.counts.done));
  localStorage.setItem("json:"+SHEET+":where",where);
  drawSaved();
}
// 「最後に JSON へ退避したのはいつか」を常に見せる。前回保存より判定が増えて
// いれば色を変える(その差分はブラウザの中にしか無い)。
function drawSaved(t){
  const el=document.getElementById("jsonstat"); if(!el) return;
  const iso=localStorage.getItem("json:"+SHEET), done=(t||tally()).done;
  if(!iso){
    el.textContent="JSON 未保存";
    el.style.color=done?"var(--exc)":"var(--muted)";
    el.title="S キーで進捗を JSON に書き出す（localStorage が消えたときの唯一の復帰手段）";
    return;
  }
  const n=+(localStorage.getItem("json:"+SHEET+":n")||0);
  const d=new Date(iso);
  el.textContent="JSON "+d.toLocaleString("ja-JP",{month:"2-digit",day:"2-digit",
    hour:"2-digit",minute:"2-digit"})+(done>n?"（+"+(done-n)+"）":"");
  el.style.color=done>n?"var(--uns)":"var(--muted)";
  el.title=(localStorage.getItem("json:"+SHEET+":where")||"")
    +" / 保存時 "+n+"件"+(done>n?" → 未保存 "+(done-n)+"件":"");
}

let toastT=null;
function toast(msg,bad){
  const el=document.getElementById("toast");
  el.textContent=msg; el.className=bad?"on bad":"on";
  clearTimeout(toastT); toastT=setTimeout(()=>{el.className=""},6000);
}

// 判定を置き換える操作は取り返しがつきにくい。必ず内訳を見せてから実行する。
function dialog(html,buttons){
  return new Promise(res=>{
    const wrap=document.getElementById("dlg"), box=document.getElementById("dlgbox");
    box.innerHTML=html+'<div class="btns"></div>';
    const bar=box.querySelector(".btns");
    const close=v=>{wrap.classList.remove("on");
      document.removeEventListener("keydown",onk,true); res(v)};
    buttons.forEach(b=>{const el=document.createElement("button");
      el.textContent=b.label; if(b.primary)el.className="primary";
      el.onclick=()=>close(b.value); bar.appendChild(el);});
    const onk=e=>{if(e.key==="Escape"){e.stopPropagation();e.preventDefault();close(null)}};
    document.addEventListener("keydown",onk,true);
    wrap.classList.add("on");
    const first=bar.querySelector("button.primary")||bar.firstChild;
    if(first)first.focus();
  });
}
function modalOn(){return document.getElementById("dlg").classList.contains("on")}

async function saveJSON(){
  const p=buildProgress(), text=JSON.stringify(p,null,1);
  let dir=null;
  if(FS_OK){
    try{dir=await getDir(true)}
    catch(e){ dir=null;
      if(e&&e.name==="AbortError") toast("フォルダを選ばなかったので、いつものダウンロードに切り替える",true);
      else toast("フォルダを開けない("+(e.message||e.name)+")のでダウンロードにする",true); }
  }
  if(dir){
    try{
      const fh=await dir.getFileHandle(PROGRESS_FILE,{create:true});
      const w=await fh.createWritable(); await w.write(text); await w.close();
      noteSaved(p,dir.name+"/"+PROGRESS_FILE);
      toast("保存した: "+dir.name+"/"+PROGRESS_FILE+"（判定済 "+p.counts.done+" 件）");
      return;
    }catch(e){ toast("フォルダへ書けない("+(e.message||e.name)+")のでダウンロードにする",true) }
  }
  download(text,PROGRESS_FILE,"application/json");
  noteSaved(p,"（ダウンロード）"+PROGRESS_FILE);
  toast("書き出した: "+PROGRESS_FILE+"（判定済 "+p.counts.done
       +" 件）ダウンロードフォルダから screening/json/ へ移すこと");
}

// 受け取った JSON は「他人が書いたかもしれないファイル」として扱う。
// 統制語彙の外・担当外ID・別シートを黙って取り込むと、xlsx へ反映する時ではなく
// κ を集計する時に初めて壊れていると分かる、という最悪の順番になる。
function parseProgress(text){
  let p; try{p=JSON.parse(text)}catch(e){return{err:"JSON として読めない: "+esc(e.message)}}
  if(!p||typeof p!=="object") return{err:"JSON の中身がオブジェクトでない"};
  if(p.format!==PROGRESS_FORMAT)
    return{err:'format が "'+PROGRESS_FORMAT+'" でない（この閲覧アプリの進捗ファイルではない）'};
  if(+p.version>PROGRESS_VERSION)
    return{err:"version "+esc(String(p.version))+" はこの閲覧アプリ(v"+PROGRESS_VERSION
              +")より新しい。make_review_app.py で HTML を作り直すこと"};
  if(p.sheet!==SHEET)
    return{err:"担当シートが違う（ファイル: "+esc(String(p.sheet))+" / このアプリ: "+SHEET
              +"）。他人の判定を自分のシートへ入れると二重スクリーニングの独立性が壊れる"};
  const dec=p.decisions;
  if(!dec||typeof dec!=="object") return{err:"decisions が無い"};
  const known={}; RECORDS.forEach(r=>known[r.record_id]=1);
  const ok={}, bad=[], unknown=[], noteless=[];
  for(const id of Object.keys(dec)){
    const s=dec[id]||{}, d=String(s.decision||""), rs=String(s.reason||""), nt=String(s.note||"");
    if(!known[id]){unknown.push(id);continue}
    if(DECISIONS.indexOf(d)<0){bad.push(id+": 判定が不正 "+JSON.stringify(s.decision));continue}
    if(d==="Exclude"){
      if(!rs){bad.push(id+": Exclude なのに除外理由が空");continue}
      if(!REASONS.some(x=>x.value===rs)){
        bad.push(id+": 統制語彙の外の除外理由 "+JSON.stringify(rs));continue}
      if(rs===REASON_OTHER&&!nt) noteless.push(id);
    }
    ok[id]={decision:d,reason:(d==="Exclude"?rs:""),note:nt};
  }
  if(bad.length) return{err:"不正な判定が "+bad.length+" 件あるので読み込まない:<br>"
    +bad.slice(0,8).map(esc).join("<br>")+(bad.length>8?"<br>… 他 "+(bad.length-8)+" 件":"")};
  if(!Object.keys(ok).length)
    return{err:"このシートの record_id が1件も入っていない（不明なID "+unknown.length+" 件）"};
  return{p:p,ok:ok,unknown:unknown,noteless:noteless};
}

function diffProgress(ok){
  let add=0,chg=0,same=0,lost=0;
  for(const id of Object.keys(ok)){
    const c=store[id];
    if(!c||!c.decision){add++;continue}
    if(c.decision===ok[id].decision&&(c.reason||"")===ok[id].reason
       &&(c.note||"")===ok[id].note) same++; else chg++;
  }
  for(const r of RECORDS){const c=store[r.record_id];
    if(c&&c.decision&&!ok[r.record_id]) lost++;}
  return {add:add,chg:chg,same:same,lost:lost};
}

function pickFileFallback(){
  return new Promise(res=>{
    const el=document.getElementById("fileinput");
    let done=false; const fin=v=>{if(!done){done=true;res(v)}};
    el.value=""; el.onchange=()=>fin(el.files[0]||null);
    // キャンセルでは change が来ない。ダイアログが閉じてフォーカスが戻ったら諦める。
    window.addEventListener("focus",()=>setTimeout(()=>fin(el.files[0]||null),500),{once:true});
    el.click();
  });
}

async function loadJSON(pick){
  let text=null,name="";
  if(!pick){  // 許可済みのフォルダに定位置のファイルがあれば、それを既定にする
    try{
      const dir=await getDir(false);
      if(dir){const fh=await dir.getFileHandle(PROGRESS_FILE);
        const f=await fh.getFile(); text=await f.text(); name=dir.name+"/"+f.name;}
    }catch(e){}
  }
  if(text===null){
    let f=null;
    if(FS_OK){
      try{
        const h=await window.showOpenFilePicker({id:"phase3b-json",multiple:false,
          types:[{description:"進捗JSON",accept:{"application/json":[".json"]}}]});
        f=await h[0].getFile();
      }catch(e){ if(e&&e.name==="AbortError") return; }
    }
    if(!f) f=await pickFileFallback();
    if(!f) return;
    text=await f.text(); name=f.name;
  }
  const r=parseProgress(text);
  if(r.err){
    const v=await dialog('<h2>読み込めない</h2><div class="dwarn">'+r.err+'</div>'
      +'<p class="dmsg">'+esc(name)+'</p>',
      [{label:"別のファイルを選ぶ",value:"other"},{label:"閉じる",value:null,primary:true}]);
    if(v==="other") return loadJSON(true);
    return;
  }
  const p=r.p, d=diffProgress(r.ok), warn=[];
  if(p.fingerprint!==FINGERPRINT)
    warn.push("指紋が違う（ファイル: "+esc(String(p.fingerprint))+" / このアプリ: "+FINGERPRINT
             +"）。判定対象の集合が違う可能性がある");
  if(r.unknown.length)
    warn.push("このシートに無い record_id を "+r.unknown.length+" 件、読み飛ばす");
  if(r.noteless.length)
    warn.push("除外理由が「"+esc(REASON_OTHER)+"」なのにメモが空: "+r.noteless.length
             +" 件（xlsx へ反映する前にメモを入れること）");
  if(d.lost)
    warn.push("<b>いま手元にある判定 "+d.lost+" 件が消える</b>（ファイルに入っていないため）");
  const saved=p.saved_at?new Date(p.saved_at).toLocaleString("ja-JP"):"不明";
  const html='<h2>進捗を読み込む</h2>'
    +'<p class="dmsg">'+esc(name)+'<br>保存 '+esc(saved)+' ／ ファイルの判定済 '
    +Object.keys(r.ok).length+' 件</p>'
    +'<table class="dtab">'
    +'<tr><td>新しく入る</td><td>'+d.add+'</td></tr>'
    +'<tr><td>上書きされる</td><td>'+d.chg+'</td></tr>'
    +'<tr><td>変わらない</td><td>'+d.same+'</td></tr>'
    +'<tr><td>消える（手元にのみある判定）</td><td>'+d.lost+'</td></tr></table>'
    +(warn.length?'<div class="dwarn">'+warn.map(w=>"・"+w).join("<br>")+'</div>':'')
    +'<p class="dmsg">読み込むと現在の判定は<b>ファイルの内容で置き換わる</b>（統合はしない）。'
    +'直前の状態は localStorage の <code>'+KEY+':backup</code> に退避する。</p>';
  const ans=await dialog(html,[{label:"置き換える",value:"ok",primary:true},
                               {label:"別のファイルを選ぶ",value:"other"},
                               {label:"やめる",value:null}]);
  if(ans==="other") return loadJSON(true);
  if(ans!=="ok") return;
  localStorage.setItem(KEY+":backup",
    JSON.stringify({saved_at:new Date().toISOString(),store:store}));
  const keep=RECORDS[view[cur]].record_id;
  store={}; Object.keys(r.ok).forEach(id=>{store[id]=r.ok[id]});
  undo=[];  // 取り消し履歴は捨てた状態を指しているので持ち越さない
  save(); rebuild(keep);
  toast("読み込んだ: "+name+"（判定済 "+tally().done+" 件）");
}

const KEYMAP=[
  ["i","Include"],["u","Unsure"],
  ...REASONS.map((x,n)=>[String(n+1),"Exclude — "+x.value]),
  ["j / →","次へ"],["k / ←","前へ"],["m","メモ欄へ"],["z","直前を取り消し"],
  ["h","ハイライト切替"],["t","訳の表示切替（原文のみ → 対訳 → 訳のみ）"],
  ["r","キー凡例の表示切替"],["e","CSV書き出し"],
  ["s","途中保存（進捗を JSON に書き出す）"],
  ["l","途中経過を読み込む（Shift+L でファイルを選ぶ）"],["?","このヘルプ"],
];
// 訳が1件でも載っていれば切替ボタンを出す。
if(RECORDS.some(r=>r.ja)){
  const b=document.getElementById("jamode");
  b.hidden=false; b.textContent="訳 "+JA_MODE_LABEL[jaMode];
  b.onclick=()=>{jaMode=(jaMode+1)%3;b.textContent="訳 "+JA_MODE_LABEL[jaMode];draw()};
}
document.getElementById("keys").innerHTML=
  ['<kbd>i</kbd> Include','<kbd>u</kbd> Unsure']
  .concat(REASONS.map((x,n)=>
    '<span class="rchip"><kbd>'+(n+1)+'</kbd> '+esc(x.value)+'</span>'))
  .concat(['<kbd>j</kbd>/<kbd>k</kbd> 移動','<kbd>m</kbd> メモ','<kbd>z</kbd> 取消',
   '<kbd>t</kbd> 訳','<kbd>r</kbd> 凡例','<kbd>e</kbd> CSV',
   '<kbd>s</kbd> JSON保存','<kbd>l</kbd> JSON読込','<kbd>?</kbd> ヘルプ'])
  .join("　");

// ---- 常時表示のキー凡例 ----
// 1-9 の並びは語彙の表示順(凍結済み)であり、Rev.24 の優先順位とは一致しない。
// 取り違えを防ぐため、各行に優先順位を併記する。
function rank(v){const i=PRIORITY.indexOf(v);return i<0?"":"優先"+(i+1);}
document.getElementById("legend").innerHTML=
  '<h3>Exclude ＋除外理由</h3>'
  +REASONS.map((x,n)=>
    '<div class="row"><kbd>'+(n+1)+'</kbd><span class="lb">'+esc(x.value)
    +'</span><span class="pr">'+rank(x.value)+'</span></div>').join("")
  +'<h3 class="sec">判定・移動</h3>'
  +[["i","Include"],["u","Unsure"],["j / →","次へ"],["k / ←","前へ"],
    ["m","メモ欄へ"],["z","直前を取り消し"],
    ["h","ハイライト切替"],["t","訳の切替"],
    ["r","この凡例を閉じる"],["e","CSV書き出し"],
    ["s","途中保存(JSON)"],["l","途中経過の読込"],
    ["?","ヘルプ"]]
   .map(([k,v])=>'<div class="row"><kbd>'+esc(k)+'</kbd><span class="lb">'+esc(v)+'</span></div>').join("")
  +'<div class="foot"><b>優先</b> は理由が2つ以上当てはまるときに'
  +'どれを選ぶかの順位（Rev.24）。'
  +'<b>数字キーの並びとは一致しません</b>'
  +'（キーは語彙の表示順）。</div>';

let legendOn=localStorage.getItem("legend:"+SHEET)!=="0";
function drawLegend(){document.body.classList.toggle("legend-on",legendOn);
  document.getElementById("lgd").textContent=(legendOn?"凡例を閉じる":"凡例 (R)");}
function toggleLegend(){legendOn=!legendOn;
  localStorage.setItem("legend:"+SHEET,legendOn?"1":"0");drawLegend();}
document.getElementById("lgd").onclick=toggleLegend;
drawLegend();
document.getElementById("helptable").innerHTML=
  KEYMAP.map(([k,v])=>"<tr><td><kbd>"+k+"</kbd></td><td>"+esc(v)+"</td></tr>").join("")
  +REASONS.map((x,n)=>"<tr><td><kbd>"+(n+1)+"</kbd></td><td><b>"+esc(x.value)+"</b><br><span style='color:var(--muted)'>"+esc(x.desc)+"</span></td></tr>").join("");

document.addEventListener("keydown",e=>{
  if(e.target.tagName==="INPUT"||e.target.tagName==="SELECT"){
    if(e.key==="Escape"||e.key==="Enter"){e.target.blur();e.preventDefault();}
    return;
  }
  if(e.ctrlKey||e.metaKey||e.altKey) return;
  if(modalOn()) return;   // 確認ダイアログを開いている間は判定キーを効かせない
  const k=e.key;
  if(k==="r"||k==="R"){toggleLegend();e.preventDefault();return}
  if(k==="?"){document.getElementById("help").classList.toggle("on");e.preventDefault();return}
  if(k==="Escape"){document.getElementById("help").classList.remove("on");return}
  if(k==="i"||k==="I"){set("Include");e.preventDefault();return}
  if(k==="u"||k==="U"){set("Unsure");e.preventDefault();return}
  if(/^[1-9]$/.test(k)){const n=+k-1; if(n<REASONS.length){set("Exclude",REASONS[n].value);e.preventDefault();}return}
  if(k==="j"||k==="ArrowRight"||k===" "){next();e.preventDefault();return}
  if(k==="k"||k==="ArrowLeft"){prev();e.preventDefault();return}
  if(k==="m"||k==="M"){document.getElementById("note").focus();e.preventDefault();return}
  if(k==="h"||k==="H"){hlOn=!hlOn;draw();return}
  if(k==="t"||k==="T"){jaMode=(jaMode+1)%3;draw();
    const b=document.getElementById("jamode"); if(b)b.textContent="訳 "+JA_MODE_LABEL[jaMode];
    return}
  if(k==="e"||k==="E"){exportCSV();return}
  if(k==="s"||k==="S"){saveJSON();e.preventDefault();return}
  if(k==="l"){loadJSON(false);e.preventDefault();return}
  if(k==="L"){loadJSON(true);e.preventDefault();return}
  if(k==="z"||k==="Z"){const u=undo.pop(); if(u){ if(u.prev&&u.prev.decision)store[u.id]=u.prev; else delete store[u.id];
    save(); rebuild(u.id);} return}
});
document.getElementById("exp").onclick=exportCSV;
document.getElementById("sav").onclick=()=>saveJSON();
document.getElementById("ld").onclick=()=>loadJSON(false);
document.getElementById("hlp").onclick=()=>document.getElementById("help").classList.toggle("on");
document.getElementById("filter").onchange=()=>{cur=0;rebuild()};
window.addEventListener("beforeunload",e=>{
  const any=Object.keys(store).length; if(any){e.preventDefault();e.returnValue=""}
});
rebuild();
</script></body></html>
"""


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--id", default="author", help="シートID (既定: author)")
    ap.add_argument("--rules", type=Path,
                    help='キーワードルールJSON: [{"label":"臨床?","any":["patient","surgery"]}] '
                         "— ハイライトと並べ替えにのみ使う。判定は書き込まない")
    ap.add_argument("--translations", type=Path,
                    help="scripts/translate_abstracts.py が作る日本語訳キャッシュ JSON。"
                         "指定すると原文と並べて表示する（訳文だけの表示は既定にしない）")
    ap.add_argument("--out", type=Path, help="出力先HTML")
    args = ap.parse_args()

    rules = []
    if args.rules:
        rules = json.loads(args.rules.read_text(encoding="utf-8"))
        if not isinstance(rules, list):
            sys.exit("[ERROR] --rules は配列であること")

    translations = {}
    if args.translations:
        if not args.translations.exists():
            sys.exit(f"[ERROR] {args.translations} が無い。先に translate_abstracts.py を実行すること。")
        translations = json.loads(args.translations.read_text(encoding="utf-8"))

    rows = load_records(args.id)
    payload = build_payload(rows, translations)
    out = args.out or (SCREENING / f"review_{args.id}.html")
    out.write_text(render(args.id, payload, rules), encoding="utf-8")

    ndone = sum(1 for r in rows if (r.get("decision") or "").strip())
    ncal = sum(1 for r in rows if r.get("calibration") == "Y")
    nja = sum(1 for r in payload if r.get("ja"))
    print(f"[OK] {out}")
    print(f"     {len(rows):,} 件（校正セット {ncal:,} 件 / 既存の判定 {ndone:,} 件）")
    if args.translations:
        stale = len(translations) - nja
        print(f"     日本語訳 {nja:,} 件を対訳表示"
              + (f"（原文と不一致で不採用 {stale:,} 件）" if stale > 0 else ""))
    print(f"     指紋 {fingerprint(payload)}（進捗JSON の照合に使う）")
    print(f"     ブラウザで開いて判定 → E キーで decisions_{args.id}.csv を書き出す")
    print(f"     途中保存: S キーで JSON 書き出し（保存先フォルダは初回に screening/json を選ぶ）")
    print(f"     途中復帰: L キーで JSON 読み込み（Shift+L でファイルを選ぶ）")
    print(f"     反映: python -X utf8 scripts/apply_review_decisions.py "
          f"--id {args.id} --input decisions_{args.id}.csv")
    print(f"           （JSON をそのまま渡してもよい: --input screening/json/progress_{args.id}.json）")


if __name__ == "__main__":
    main()
