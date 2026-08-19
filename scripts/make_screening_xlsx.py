#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_screening_xlsx.py — Phase 3b 判定シートの Excel 版を生成する
================================================================================

【何を】
`make_screening_sheets.py` が作った CSV 判定シートから、評価者が実際に作業しやすい
`.xlsx` を生成する。CSV は機械可読の正、xlsx は**人が記入するための作業ファイル**。

【なぜ xlsx にするか】
1,700件以上を CSV エディタで読み書きするのは現実的でない。Excel なら
  - decision 列と reason 列を**ドロップダウン**にでき、表記ゆれが発生しない
    (include/INCLUDE/Inc、「患者対象」/「P: 患者が対象」/「対象者が患者」…)。
    reason が統制語彙になると、除外理由の内訳をそのまま集計・報告できる
  - 要旨を折り返して読める
  - 見出し固定・フィルタで「未記入だけ表示」などができる
  - メタ列をロックして誤編集を防げる
判定の質と、後段の集計の安全性がそのまま上がる。

【アクセシビリティで気をつけたこと】
- **色だけに意味を持たせない**。要旨欠落は塗り分けだけでなく `has_abstract` 列の
  "N" という文字でも判別でき、オートフィルタでも絞り込める
- 前景色と背景色のコントラストを確保し、淡い塗りに濃い文字を組み合わせる
- 既定フォントサイズを 11pt、本文列は折り返し + 上揃えで行を読みやすく
- 見出し行を固定(freeze panes)し、スクロールしても列の意味を見失わない
- 記入列だけロックを外し、**Tab キーで decision → reason → note と移動できる**
- セル結合を使わない(スクリーンリーダーと並べ替えの両方で問題になるため)
- 列見出しに日本語の説明を入れ、別紙を見なくても意味が分かるようにする

【入出力】
  入力: screening/sheet_<id>.csv(+ 記入済みがあればその内容を引き継ぐ)
  出力: screening/sheet_<id>.xlsx

実行:
  python -X utf8 scripts/make_screening_xlsx.py
  python -X utf8 scripts/make_screening_xlsx.py --only author
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

_HERE = Path(__file__).resolve().parent
ROOT = _HERE.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(_HERE))

csv.field_size_limit(10 ** 9)

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("[ERROR] openpyxl が必要です:  pip install openpyxl")

from make_screening_sheets import REVIEWERS  # noqa: E402

# --- 配色 -------------------------------------------------------------------
# 淡い背景 × 濃い文字でコントラストを確保する。色は補助であり、
# 同じ情報が必ず文字(列の値)でも取れるようにしてある。
C_HEADER_BG = "1F3864"   # 濃紺
C_HEADER_FG = "FFFFFF"
C_INPUT_BG = "FFF7E6"    # 記入列(淡い橙)
C_NOABS_BG = "FDE7E9"    # 要旨欠落行(淡い赤)
C_BORDER = "BFBFBF"
C_NOTE = "44546A"

# 列定義: (CSV列名, 表示名, 幅, 折り返すか, 記入列か)
COLUMNS = [
    ("record_id",    "ID",              13, False, False),
    ("decision",     "判定 ★",          13, False, True),
    ("reason",       "除外理由 ★",      34, True,  True),
    ("note",         "メモ",            22, True,  True),
    ("title",        "タイトル",         52, True,  False),
    ("abstract",     "要旨",             95, True,  False),
    ("venue",        "掲載先",           30, True,  False),
    ("year",         "年",                7, False, False),
    ("rank",         "ランク",           11, False, False),
    ("has_abstract", "要旨有無",          9, False, False),
    ("source",       "取得経路",         12, False, False),
    ("calibration",  "校正セット",       11, False, False),
    ("abstract_source", "要旨の出所",    12, False, False),
    ("kw_groups",    "概念群",            8, False, False),
    ("doi",          "DOI",              26, False, False),
    ("block",        "ブロック",          9, False, False),
]

HEADER_HELP = {
    "判定 ★": "この列を埋めてください。セルを選ぶとドロップダウンが出ます。\n"
              "Include=残す / Exclude=除外 / Unsure=判断保留(協議に回ります)",
    "除外理由 ★": "Exclude のときは必須です。ドロップダウンから1つ選んでください。\n"
                  "選択肢の定義は「はじめに」シートにあります。\n"
                  "複数該当するときは最も明白なものを1つ選び、残りはメモへ。\n"
                  "「その他」を選んだときはメモに理由を必ず書いてください。",
    "メモ": "任意。協議で持ち出したい論点があれば。\n"
            "ただし除外理由が「その他」のときは必須です。",
    "要旨有無": "N = 要旨が無く、タイトルだけで判断することになる文献です。",
    "取得経路": "database = データベース検索で見つけたもの\n"
                "snowballing = 引用探索(参考文献・被引用)で見つけたもの\n"
                "判定基準はどちらも同じです。PRISMA の報告で区別するための記録です。",
    "校正セット": "Y = 3名全員が判定する校正セット(評価者間一致度 κ の算出に使う)\n"
                  "N = 著者がまず判定し、Exclude になったものだけ第2評価者が確認する\n"
                  "※ 判定基準はどちらも同じです。",
    "要旨の出所": "database = 検索結果に元から含まれていた要旨\n"
                  "enriched = 要旨が無かったため DOI から外部で補完したもの\n"
                  "none = 要旨を取得できず、タイトルのみで判断することになる\n"
                  "※ enriched も判定材料としては同じように使ってください。",
    "概念群": "検索クエリの3概念群がいくつ当たったか(0〜3)。\n"
              "**読む順序の目安にすぎません。この値で判定しないでください。**",
    "ランク": "Phase 2 で既に基準を満たしています。判定材料にはしないでください。",
}

DECISIONS = ["Include", "Exclude", "Unsure"]

# --- 除外理由の統制語彙 -------------------------------------------------------
# なぜ自由記述をやめたか:
#   (1) PRISMA 2020 は除外理由の記録を求めるが、自由記述だと後から**数えられない**。
#       統制語彙にしておけば「理由別の除外件数」をそのまま表に載せられる。
#   (2) 評価者3名で表記が割れると、同じ理由が別物として集計される。
#   (3) 基準名で選ばせること自体が「PICOS のどれに抵触したか」を毎回考えさせる。
#
# なぜ C(比較対象)が無いか:
#   比較条件は要旨に書かれないことが多く、Title/Abstract 段で C を根拠に落とすと
#   誤除外(回復不能)を招く。C 抵触の判断は Phase 4(全文)に送る。
#
# なぜ「その他」があるか:
#   閉じた語彙は、当てはまらない事例を無理に既存カテゴリへ押し込ませる。
#   逃げ道を用意したうえで**メモへの記述を必須**にすれば、事後に語彙を見直せる。
#
# (選択肢, 定義) — 選択肢の文字列がそのままセルの値になる。
# 変更するときは EXAMPLE(make_screening_example.py)と screening_protocol.md も直すこと。
EXCLUDE_REASONS = [
    ("P: 対象者が不適合",
     "健常成人でない(小児・高齢者・患者・特定の専門職集団など)"),
    ("I: HMD-VR でない",
     "デスクトップ画面・AR/MR・CAVE・実環境のみなど、HMD を用いた VR でない"),
    ("I: スケール操作/多感覚刺激が無い",
     "身体・空間スケールの操作も、多感覚刺激の提示も行っていない"),
    ("O: スケール知覚の測定が無い",
     "自己・環境のスケール変容を測る定量指標が無い"),
    ("S: ユーザー実験が無い",
     "技術提案・デモ・シミュレーションのみで、実証的なユーザー実験を伴わない"),
    ("S: 原著論文でない",
     "総説・サーベイ・抄録のみ・ポスター・基調講演・書籍など"),
    ("スコープ外(主題が無関係)",
     "基準を当てるまでもなく、本レビューの主題と無関係"),
    ("重複(同一研究の別報告)",
     "既に対象に含まれている研究の別報告・拡張版・重複登録"),
    ("その他",
     "上のどれにも当てはまらない。※メモ列に理由を必ず書くこと"),
]
REASON_VALUES = [v for v, _ in EXCLUDE_REASONS]
REASON_OTHER = "その他"

# ドロップダウンの選択肢を置く隠しシート。
# インラインの選択肢リスト('"a,b,c"')は Excel の仕様で 255 文字までしか入らず、
# 日本語の選択肢では即座に溢れる。シート上の範囲を参照する形なら制限が無い。
LIST_SHEET = "_選択肢"

INTRO = [
    ("Phase 3b  Title/Abstract スクリーニング  判定シート", "title"),
    ("", ""),
    ("担当: {reviewer}", "lead"),
    ("担当件数: {n:,} 件", "lead"),
    ("", ""),
    ("■ やること", "h"),
    ("「判定」シートの ★ が付いた2列を埋めてください。どちらも選択式です。", ""),
    ("  ・判定 … Include(残す) / Exclude(除外) / Unsure(保留) から選ぶ", ""),
    ("  ・除外理由 … Exclude のときは必須。下の一覧から1つ選ぶ", ""),
    ("", ""),
    ("■ 除外理由の選択肢", "h"),
    ("集計して「理由別の除外件数」を論文に載せるため、選択式にしてあります。", ""),
    ("__REASON_TABLE__", ""),
    ("複数該当するときは、最も明白なものを1つ選び、残りはメモに書いてください。", ""),
    ("当てはまるものが無ければ「その他」を選び、メモに理由を必ず書いてください。", "lead"),
    ("(C(比較対象)は選択肢にありません。比較条件は要旨に書かれないことが多く、", ""),
    ("  この段階で C を根拠に落とすと誤除外になるためです。全文評価に送ります。)", ""),
    ("", ""),
    ("■ 判定の考え方", "h"),
    ("除外できると確信できないものは残してください(Include)。", ""),
    ("全文を読めば分かることを、この段階で切らないためです。", ""),
    ("迷ったら Unsure で構いません。Unsure は後で協議にかけます。", ""),
    ("", ""),
    ("■ お願い", "h"),
    ("他の評価者のファイルは開かないでください。", "warn"),
    ("二重スクリーニングは互いの判定を見ないことが前提で、", ""),
    ("見てしまうと評価者間一致度(κ)が意味を失います。", ""),
    ("", ""),
    ("■ 使い方のヒント", "h"),
    ("・見出し行は固定してあります。横スクロールしても列名が残ります。", ""),
    ("・オートフィルタで「判定」を (空白) で絞ると未記入だけ表示できます。", ""),
    ("・薄い赤の行は要旨が無い文献です(「要旨有無」列が N)。", ""),
    ("  タイトルだけで判断することになるので、無理なら Unsure にしてください。", ""),
    ("・「要旨の出所」列が enriched のものは、検索結果に要旨が無かったため", ""),
    ("  DOI から外部で補完したものです。判定材料としては同じように使ってください。", ""),
    ("・「取得経路」列は database / snowballing の2種類があります。", ""),
    ("  判定基準はどちらも同じです。区別は PRISMA の報告に使うだけです。", ""),
    ("・「概念群」は読む順序の目安です。並べ替えの基準に使ってあるだけで、", ""),
    ("  この数値で判定しないでください。", ""),
    ("・入力する3列以外はロックしてあります(誤編集の防止)。", ""),
    ("", ""),
    ("■ 要旨を全文読むには", "h"),
    ("要旨は平均1,300字ほどあり、セルには冒頭の5行程度しか表示されません。", ""),
    ("全文を読むときは次のどちらかで開いてください。", ""),
    ("  (1) その行の行番号を右クリック →「行の高さの自動調整」", "lead"),
    ("      読み終わったら元に戻せます。行の高さ変更は許可してあります。", ""),
    ("  (2) 要旨のセルを選び、数式バー右端の∨ボタンで数式バーを広げる", "lead"),
    ("行を全部広げると走査しづらくなるので、既定は詰めた高さにしてあります。", ""),
    ("", ""),
    ("■ 終わったら", "h"),
    ("このファイルを著者に渡してください。集計とκの算出はスクリプトが行います。", ""),
]


def load_rows(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="", errors="replace") as f:
        return list(csv.DictReader(f))


def build_intro(ws, reviewer: str, n: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 104
    styles = {
        "title": Font(size=16, bold=True, color=C_HEADER_BG),
        "h":     Font(size=12, bold=True, color=C_HEADER_BG),
        "lead":  Font(size=11, bold=True),
        "warn":  Font(size=11, bold=True, color="9C0006"),
        "":      Font(size=11),
    }
    # 語彙表はプレースホルダを実際の選択肢に展開する(EXCLUDE_REASONS と二重管理しない)
    lines: list[tuple[str, str]] = []
    for text, kind in INTRO:
        if text == "__REASON_TABLE__":
            lines += [(f"  ・{v}  … {d}", "") for v, d in EXCLUDE_REASONS]
        else:
            lines.append((text, kind))

    for i, (text, kind) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1,
                    value=text.format(reviewer=reviewer, n=n) if text else "")
        c.font = styles[kind]
        c.alignment = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[i].height = 24 if kind in ("title", "h") else 18
    ws.protection.sheet = True   # 説明シートは編集不可


def build_reason_list_sheet(wb) -> str:
    """除外理由の選択肢を置く隠しシートを用意し、DataValidation 用の参照文字列を返す。

    インラインの選択肢リストは Excel の 255 文字制限に掛かるため、範囲参照にする。
    定義列も並べて置いておくと、シートを表示すれば語彙表としても読める。
    """
    if LIST_SHEET in wb.sheetnames:
        ws = wb[LIST_SHEET]
    else:
        ws = wb.create_sheet(LIST_SHEET)
        for i, (value, desc) in enumerate(EXCLUDE_REASONS, start=1):
            ws.cell(row=i, column=1, value=value)
            ws.cell(row=i, column=2, value=desc)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 60
        ws.protection.sheet = True
        ws.sheet_state = "hidden"     # 作業の邪魔にならないよう既定では隠す
    return f"={LIST_SHEET}!$A$1:$A${len(EXCLUDE_REASONS)}"


def build_sheet(ws, rows: list[dict]) -> None:
    thin = Side(style="thin", color=C_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- 見出し ---
    for j, (_key, label, width, _wrap, _inp) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(bold=True, color=C_HEADER_FG, size=11)
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(j)].width = width
        if label in HEADER_HELP:
            c.comment = Comment(HEADER_HELP[label], "screening protocol")
            c.comment.width = 320
            c.comment.height = 110
    ws.row_dimensions[1].height = 32

    # --- 本体 ---
    input_fill = PatternFill("solid", fgColor=C_INPUT_BG)
    for i, row in enumerate(rows, start=2):
        for j, (key, _label, _w, wrap, is_input) in enumerate(COLUMNS, start=1):
            val = row.get(key, "")
            if key in ("year", "kw_groups", "block"):
                try:
                    val = int(val)
                except (TypeError, ValueError):
                    pass
            c = ws.cell(row=i, column=j, value=val)
            c.alignment = Alignment(
                vertical="top", wrap_text=wrap,
                horizontal="center" if key in ("year", "kw_groups", "block",
                                               "has_abstract", "decision") else "left")
            c.border = border
            c.font = Font(size=11)
            if is_input:
                c.fill = input_fill
                c.protection = Protection(locked=False)
            else:
                c.protection = Protection(locked=True)
        # 要旨は中央値 1,278 文字あり、セルに全文を収めると1行が15行分の高さになって
        # 走査できなくなる。ここでは「タイトル全文 + 要旨の冒頭5行」が見える高さに抑え、
        # 全文を読む手順を「はじめに」シートに明示する(行の書式変更は保護から除外済み)。
        ws.row_dimensions[i].height = 85

    last = len(rows) + 1
    ws.freeze_panes = "B2"                     # 見出し行 + ID 列を固定
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"

    # --- 判定列のドロップダウン ---
    dcol = get_column_letter(1 + [c[0] for c in COLUMNS].index("decision"))
    dv = DataValidation(type="list", formula1='"' + ",".join(DECISIONS) + '"',
                        allow_blank=True, showDropDown=False)
    dv.error = "Include / Exclude / Unsure のいずれかを選んでください。"
    dv.errorTitle = "入力できない値です"
    dv.prompt = "Include=残す / Exclude=除外 / Unsure=保留"
    dv.promptTitle = "判定"
    ws.add_data_validation(dv)
    dv.add(f"{dcol}2:{dcol}{last}")

    # --- 除外理由列のドロップダウン -----------------------------------------
    # 選択肢が長いのでインラインリスト(255文字上限)は使えない。隠しシートの範囲を参照する。
    rcol = get_column_letter(1 + [c[0] for c in COLUMNS].index("reason"))
    ref = build_reason_list_sheet(ws.parent)
    rv = DataValidation(type="list", formula1=ref, allow_blank=True, showDropDown=False)
    rv.error = ("一覧から選んでください。当てはまるものが無ければ「その他」を選び、"
                "メモ列に理由を書いてください。")
    rv.errorTitle = "入力できない値です"
    rv.prompt = "Exclude のときは必須。抵触した基準を1つ選ぶ(定義は「はじめに」シート)"
    rv.promptTitle = "除外理由"
    ws.add_data_validation(rv)
    rv.add(f"{rcol}2:{rcol}{last}")

    # --- 条件付き書式 -------------------------------------------------------
    # (1) 要旨が無い行を淡く塗る。※色は補助で、「要旨有無」列の N でも判別できる
    hcol = get_column_letter(1 + [c[0] for c in COLUMNS].index("has_abstract"))
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(len(COLUMNS))}{last}",
        FormulaRule(formula=[f'${hcol}2="N"'],
                    fill=PatternFill("solid", fgColor=C_NOABS_BG), stopIfTrue=False))
    # (2) Exclude なのに理由が空のセルを目立たせる(提出前の自己チェック用)
    ws.conditional_formatting.add(
        f"{rcol}2:{rcol}{last}",
        FormulaRule(formula=[f'AND(${dcol}2="Exclude",LEN(TRIM(${rcol}2))=0)'],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                    font=Font(color="9C0006", bold=True), stopIfTrue=False))
    # (3) 理由が「その他」なのにメモが空のセルを目立たせる。
    #     統制語彙の逃げ道は、記述が伴わないと事後に語彙を見直せず意味を失う。
    ncol = get_column_letter(1 + [c[0] for c in COLUMNS].index("note"))
    ws.conditional_formatting.add(
        f"{ncol}2:{ncol}{last}",
        FormulaRule(formula=[f'AND(${rcol}2="{REASON_OTHER}",LEN(TRIM(${ncol}2))=0)'],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                    font=Font(color="9C0006", bold=True), stopIfTrue=False))

    # 記入列以外をロック(パスワードなし。誤操作の防止が目的で秘匿ではない)
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def main() -> None:
    ap = argparse.ArgumentParser(description="Phase 3b 判定シートの Excel 版を生成")
    ap.add_argument("--dir", type=Path, default=ROOT / "screening")
    ap.add_argument("--prefix", type=str, default="",
                    help="シート名の接頭辞(stage2_ など)")
    ap.add_argument("--only", type=str, default="",
                    help="特定の評価者だけ生成(author / kataoka / watanabe)")
    ap.add_argument("--force", action="store_true",
                    help="既存の .xlsx を上書きする(記入済みの内容は失われる)")
    args = ap.parse_args()

    targets = [args.only] if args.only else list(REVIEWERS)
    for rev in targets:
        if rev not in REVIEWERS:
            sys.exit(f"[ERROR] 未知の評価者: {rev}(有効: {', '.join(REVIEWERS)})")
        src = args.dir / f"{args.prefix}sheet_{rev}.csv"
        dst = args.dir / f"{args.prefix}sheet_{rev}.xlsx"
        if not src.exists():
            print(f"[SKIP] {src.name} が無い。先に make_screening_sheets.py を実行すること")
            continue
        if dst.exists() and not args.force:
            print(f"[SKIP] {dst.name} は既にある(記入済みを壊さないため生成しない。"
                  f"作り直すなら --force)")
            continue

        rows = load_rows(src)
        wb = Workbook()
        intro = wb.active
        intro.title = "はじめに"
        build_intro(intro, REVIEWERS[rev], len(rows))
        ws = wb.create_sheet("判定")
        build_sheet(ws, rows)
        wb.active = 1          # 開いたとき「判定」シートを表示
        wb.save(dst)

        no_abs = sum(1 for r in rows if r.get("has_abstract") == "N")
        print(f"[INFO] 出力: {dst.name}  {len(rows):,} 行  担当={REVIEWERS[rev]}"
              f"  (うち要旨なし {no_abs} 件)")

    print("\n[NEXT] 各評価者に .xlsx を配布する。記入後は screening/ に戻してもらい、")
    print("       `python -X utf8 scripts/score_screening.py` で集計する")
    print("       (xlsx があれば xlsx を、無ければ CSV を読む)。")


if __name__ == "__main__":
    main()
