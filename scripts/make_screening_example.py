#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_screening_example.py — 配布時に見せる「記入見本」シートを作る
================================================================================

【何を】
評価者に判定シートを配るとき、**どう記入すればよいかを示す見本**(1ファイル)を作る。
`screening/EXAMPLE_記入見本.xlsx`。

【★ 実際の判定は絶対に入れない】
本プロトコルは `rule.md` Rev.2 で **包含/除外判定に LLM を一切使用しない**と定めており、
原稿にも "no large language model was used at any eligibility decision point" と記載している。
したがって本スクリプトは **判定対象1,052件には一切触れない**。見本に使うのは:

  - Phase 3a で**機械的に除外済み**の文献(`step3_kw_excluded.csv`)
    … 除外理由が決定論的に記録済みなので、それを転記するだけで済む
  - gold set のうち**判定対象に含まれない**もの
    … 著者が既に「必須」と分類済みなので、新たな判定を伴わない

見本の各行には `record_id = EXAMPLE-n` を振り、**実データと取り違えようがない**ようにする。
見本シートは `assignment.csv` に載らないため、集計(`score_screening.py`)からも
構造的に除外される。

【使い方】
  python -X utf8 scripts/make_screening_example.py
  → screening/EXAMPLE_記入見本.xlsx を評価者に配る(実シートと一緒に)

**配布前に著者が内容を確認し、必要なら判定・理由を書き換えること。**
見本の判定は「記入の形式」を示すためのもので、著者の判断として提示される。
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

_HERE = Path(__file__).resolve().parent
ROOT = _HERE.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(_HERE))

csv.field_size_limit(10 ** 9)

from make_screening_sheets import SHEET_COLS, kw_group_count, norm_doi  # noqa: E402

# 見本に使う文献(すべて判定対象1,052件の**外**)。
#   src: 'excluded' = step3_kw_excluded.csv / 'gold' = self_scale_references.csv
EXAMPLES = [
    # --- 明確に Include ---
    dict(no=1, src="gold", doi="10.1145/3041164.3041204",
         decision="Include", reason="",
         note="典型的な Include。HMD 下でのサイズ知覚の実験で、PICOS の I/O/S をすべて満たす。"
              "理由欄は Include のときは空で構わない"),
    # --- 明確に Exclude(P基準) ---
    dict(no=2, src="excluded", doi="10.1109/VR46266.2020.00078",
         decision="Exclude", reason="P: 対象者が不適合",
         note="Exclude のときは**どの基準に抵触したか**をドロップダウンから選ぶ。"
              "対象が患者・外科医で健常成人でないので P。複数該当するとき"
              "(これは主題もスコープ外)は最も明白なものを1つ選び、残りはこのメモへ"),
    # --- 明確に Exclude(I基準) ---
    dict(no=3, src="excluded", doi="10.1109/TVCG.2015.2440231",
         decision="Exclude", reason="I: HMD-VR でない",
         note="デスクトップモニタ提示で HMD を用いていない。"
              "実験としては良質でも介入が HMD-VR でなければ I 基準で落ちる"),
    # --- 明確に Exclude(S基準) ---
    dict(no=4, src="excluded", doi="10.1109/VR.2017.7892233",
         decision="Exclude", reason="S: ユーザー実験が無い",
         note="システム提案・デモで、ユーザー実験による評価を伴わないもの"),
    # --- Unsure(判断保留) ---
    dict(no=5, src="excluded", doi="10.1145/3533376",
         decision="Unsure", reason="",
         note="迷ったら Unsure でよい。**無理に二択にしないこと。**"
              "Unsure は協議に回るので、判断材料が足りないときの正しい選択肢。"
              "note に迷った理由を書いておくと協議が早い"),
]

DOC_ROWS = [
    ("■ これは記入の見本です。実際の判定対象ではありません。", "warn"),
    ("", ""),
    ("各行の「メモ」列に、その判定をどう書いたかの解説を入れています。", ""),
    ("自分のシート(sheet_あなたの名前.xlsx)を開いて、同じ要領で記入してください。", ""),
    ("", ""),
    ("■ 記入するのは2列だけです", "h"),
    ("  ・判定 ★     … Include(残す) / Exclude(除外) / Unsure(保留) をドロップダウンで選ぶ", ""),
    ("  ・除外理由 ★ … Exclude のときは必須。**こちらもドロップダウンから選ぶ**", ""),
    ("", ""),
    ("■ 除外理由が選択式になっています", "h"),
    ("__REASON_TABLE__", ""),
    ("複数該当するときは最も明白なものを1つ選び、残りはメモへ(見本 EXAMPLE-2 参照)。", ""),
    ("当てはまるものが無ければ「その他」を選び、メモに理由を必ず書いてください。", "warn"),
    ("", ""),
    ("■ PICOS 基準(Title/Abstract レベルに緩めたもの)", "h"),
    ("  P 対象者   … 健常成人。小児・高齢者(臨床対象)・患者は除外", ""),
    ("  I 介入     … HMD を用いた VR 環境での身体/空間スケール操作、多感覚刺激の提示", ""),
    ("  C 比較     … スケール操作の有無、倍率間の比較、視覚単独 vs 多感覚", ""),
    ("  O 評価指標 … 自己/環境のスケール変容の定量測定", ""),
    ("  S 研究設計 … 客観的なユーザー実験にもとづく実証研究", ""),
    ("", ""),
    ("■ いちばん大事な原則", "h"),
    ("**除外できると確信できないものは残してください(Include)。**", "warn"),
    ("全文を読めば分かることを、この段階で切らないためです。", ""),
    ("この段階で誤って除外すると、その論文は二度と検討されません。", ""),
    ("逆に誤って残しても、次の全文評価で落とせるだけです。", ""),
]


def _load(path: Path, doi_col: str, title_col: str, abs_col: str,
          venue_col: str, year_col: str) -> dict[str, dict]:
    out = {}
    if not path.exists():
        return out
    with path.open(encoding="utf-8-sig", newline="", errors="replace") as f:
        for r in csv.DictReader(f):
            d = norm_doi(r.get(doi_col, ""))
            if d and d not in out:
                out[d] = {"title": (r.get(title_col) or "").strip(),
                          "abstract": (r.get(abs_col) or "").strip(),
                          "venue": (r.get(venue_col) or "").strip(),
                          "year": (r.get(year_col) or "").strip()}
    return out


def main() -> None:
    excluded = _load(ROOT / "step3_kw_excluded.csv", "DOI", "Title",
                     "Abstract Note", "Publication Title", "Publication Year")
    gold = {}
    p = ROOT / "self_scale_references.csv"
    with p.open(encoding="utf-8-sig", newline="", errors="replace") as f:
        for r in csv.DictReader(f):
            gold[norm_doi(r.get("DOI_or_URL", ""))] = {
                "title": (r.get("Title") or "").strip(), "abstract": "",
                "venue": (r.get("Venue") or "").strip(),
                "year": (r.get("Year") or "").strip()}

    # 汚染チェック: 見本の文献が判定対象に入っていないことを確認する
    apath = ROOT / "screening" / "assignment.csv"
    target_dois = set()
    if apath.exists():
        with apath.open(encoding="utf-8-sig", newline="") as f:
            target_dois = {norm_doi(r.get("doi", "")) for r in csv.DictReader(f)}

    rows = []
    for ex in EXAMPLES:
        d = norm_doi(ex["doi"])
        if d in target_dois:
            sys.exit(f"[ERROR] 見本 #{ex['no']} の DOI {d} は判定対象に含まれています。"
                     f"見本には判定対象外の文献しか使えません(汚染防止)")
        src = excluded if ex["src"] == "excluded" else gold
        rec = src.get(d)
        if not rec:
            print(f"[WARN] 見本 #{ex['no']}: {d} が見つからない。スキップ")
            continue
        rows.append({
            "record_id": f"EXAMPLE-{ex['no']}",
            "block": "—", "calibration": "—", "source": "見本",
            "kw_groups": kw_group_count(rec["title"], rec["abstract"]),
            "has_abstract": "Y" if rec["abstract"] else "N",
            "abstract_source": "—",
            "title": rec["title"], "abstract": rec["abstract"],
            "venue": rec["venue"], "year": rec["year"], "doi": d, "rank": "—",
            "decision": ex["decision"], "reason": ex["reason"], "note": ex["note"],
        })

    if not rows:
        sys.exit("[ERROR] 見本に使える文献が1件も見つかりませんでした")

    out_csv = ROOT / "screening" / "EXAMPLE_記入見本.csv"
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SHEET_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"[INFO] 出力: {out_csv.name}  ({len(rows)} 行)")

    # Excel 版
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("[WARN] openpyxl が無いため xlsx は作らない")
        return

    from make_screening_xlsx import EXCLUDE_REASONS, REASON_VALUES, build_sheet, C_HEADER_BG

    # 見本の理由が統制語彙から外れていたら、配ったシートのドロップダウンで再現できない
    for ex in EXAMPLES:
        if ex["reason"] and ex["reason"] not in REASON_VALUES:
            sys.exit(f"[ERROR] 見本 #{ex['no']} の除外理由「{ex['reason']}」は"
                     f"統制語彙に無い。make_screening_xlsx.EXCLUDE_REASONS と揃えること")

    doc_rows = []
    for text, kind in DOC_ROWS:
        if text == "__REASON_TABLE__":
            doc_rows += [(f"  ・{v}  … {d}", "") for v, d in EXCLUDE_REASONS]
        else:
            doc_rows.append((text, kind))

    wb = Workbook()
    intro = wb.active
    intro.title = "はじめに"
    intro.sheet_view.showGridLines = False
    intro.column_dimensions["A"].width = 104
    styles = {"h": Font(size=12, bold=True, color=C_HEADER_BG),
              "warn": Font(size=12, bold=True, color="9C0006"),
              "": Font(size=11)}
    intro["A1"] = "Phase 3b 判定シート  記入見本"
    intro["A1"].font = Font(size=16, bold=True, color=C_HEADER_BG)
    intro.row_dimensions[1].height = 26
    for i, (text, kind) in enumerate(doc_rows, start=3):
        c = intro.cell(row=i, column=1, value=text)
        c.font = styles[kind]
        c.alignment = Alignment(vertical="center")
        intro.row_dimensions[i].height = 22 if kind in ("h", "warn") else 18
    intro.protection.sheet = True

    ws = wb.create_sheet("記入見本")
    build_sheet(ws, rows)
    # 見本であることを視覚的にも示す(色だけに頼らず、source 列に「見本」と入れてある)
    ws["A1"].fill = PatternFill("solid", fgColor="9C0006")
    wb.active = 1
    out_xlsx = ROOT / "screening" / "EXAMPLE_記入見本.xlsx"
    wb.save(out_xlsx)
    print(f"[INFO] 出力: {out_xlsx.name}")
    print("\n[NEXT] 配布前に著者が内容を確認し、必要なら判定・理由を書き換えること。")
    print("       見本の文献はすべて判定対象1,052件の**外**から採ってある(汚染防止)。")


if __name__ == "__main__":
    main()
