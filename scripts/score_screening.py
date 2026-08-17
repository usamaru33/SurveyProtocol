#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_screening.py — Phase 3b の集計(Cohen's κ・不一致抽出・最終判定の確定)
================================================================================

【何を】
評価者ごとの判定シート(`screening/sheet_<id>.csv`)を突き合わせて、

  1. 記入の妥当性を検査する(未記入・不正値・担当外レコード・取り違え)
  2. **ペアごとの Cohen's κ とその平均**を算出する(校正セットのみ。Rev.17)
  3. 不一致と Unsure を「協議リスト」に書き出す
  4. 協議結果が記入されていれば、最終判定シートを確定する

【なぜ人手判定なのに集計をコードにするのか】
判定そのものは人が行うが、**集計を手作業でやると再現できない**。κ の値は査読で
問われる数値であり、どの文献がどう扱われたかは監査対象になる。決定論的な集計を
コード化しておけば、判定シートさえあれば第三者が同じ数値を再現できる。
LLM は使わない(rule.md Rev.2)。

【κ の報告方針】
Include / Exclude / Unsure の**3カテゴリで算出したものを主報告**とする。
あわせて、協議前に Unsure を Include とみなした2カテゴリ版も参考として出す
(本プロトコルの未解決時のルールが「Include 側に倒す」であるため)。
どちらを本文に載せるかは著者が決めるが、**両方を出したうえで選ぶ**ことで
恣意的な選択に見えないようにする。

【解決ルール(rule.md Phase 3b)】
不一致・Unsure は協議で解決する。協議で解決できない場合は **Include 側に倒して**
Phase 4(全文評価)へ送る(再現率優先)。

【入出力】
  入力: screening/assignment.csv, screening/sheet_<id>.csv
  出力: screening/consensus_worklist.csv  協議が必要なレコード(協議前)
        screening/final_decisions.csv     最終判定(協議結果が揃っている場合のみ)
        標準出力                          検査結果・κ・分布

実行:
  python -X utf8 scripts/score_screening.py
  python -X utf8 scripts/score_screening.py --strict   # 未記入があればエラー終了
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter, defaultdict
from itertools import combinations
from pathlib import Path

_HERE = Path(__file__).resolve().parent
ROOT = _HERE.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(_HERE))

csv.field_size_limit(10 ** 9)

from make_screening_sheets import REVIEWERS, second_reviewer_of  # noqa: E402

VALID = {"include", "exclude", "unsure"}
CANON = {"include": "Include", "exclude": "Exclude", "unsure": "Unsure"}


def cohen_kappa(pairs: list[tuple[str, str]]) -> tuple[float, float, float, int]:
    """(kappa, observed_agreement, expected_agreement, n) を返す。

    κ = (Po - Pe) / (1 - Pe)。Pe は各評価者のカテゴリ周辺確率の積和。
    Pe == 1 の場合(両者が全件同一カテゴリ)は κ が定義できないため nan を返す。
    """
    n = len(pairs)
    if n == 0:
        return float("nan"), float("nan"), float("nan"), 0
    cats = sorted({c for p in pairs for c in p})
    po = sum(1 for a, b in pairs if a == b) / n
    ca = Counter(a for a, _ in pairs)
    cb = Counter(b for _, b in pairs)
    pe = sum((ca[c] / n) * (cb[c] / n) for c in cats)
    if abs(1 - pe) < 1e-12:
        return float("nan"), po, pe, n
    return (po - pe) / (1 - pe), po, pe, n


def landis_koch(k: float) -> str:
    if k != k:      # nan
        return "算出不能"
    if k < 0:       return "poor(偶然以下)"
    if k < 0.21:    return "slight"
    if k < 0.41:    return "fair"
    if k < 0.61:    return "moderate"
    if k < 0.81:    return "substantial"
    return "almost perfect"


def _load_csv(path: Path) -> dict[str, dict]:
    with path.open(encoding="utf-8-sig", newline="", errors="replace") as f:
        return {r["record_id"]: r for r in csv.DictReader(f) if r.get("record_id")}


def _load_xlsx(path: Path) -> dict[str, dict]:
    """Excel 版の判定シートを読む。

    評価者が実際に記入するのは xlsx(`make_screening_xlsx.py` 製)なので、
    集計はそちらを正として読めなければならない。表示名の見出し
    (「判定」「除外理由」…)を CSV の列名(decision/reason/…)へ戻して扱う。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("[ERROR] xlsx を読むには openpyxl が必要:  pip install openpyxl")

    from make_screening_xlsx import COLUMNS
    label2key = {label: key for key, label, *_ in COLUMNS}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["判定"] if "判定" in wb.sheetnames else wb[wb.sheetnames[-1]]
    it = ws.iter_rows(values_only=True)
    header = [(str(h).strip() if h is not None else "") for h in next(it, [])]
    keys = [label2key.get(h, h) for h in header]

    out: dict[str, dict] = {}
    for values in it:
        row = {k: ("" if v is None else str(v).strip())
               for k, v in zip(keys, values)}
        rid = row.get("record_id", "")
        if rid:
            out[rid] = row
    wb.close()
    return out


def load_sheet(path: Path) -> dict[str, dict]:
    """判定シートを読む。**同名の .xlsx があればそちらを優先**する。

    評価者は xlsx で作業し、CSV は生成時の雛形として残るため、
    CSV を先に読むと「未記入」と誤判定してしまう。
    """
    xlsx = path.with_suffix(".xlsx")
    if xlsx.exists():
        return _load_xlsx(xlsx)
    if path.exists():
        return _load_csv(path)
    return {}


def main() -> None:
    ap = argparse.ArgumentParser(description="Phase 3b の集計と Cohen's κ 算出")
    ap.add_argument("--dir", type=Path, default=ROOT / "screening")
    ap.add_argument("--strict", action="store_true",
                    help="未記入・不正値があれば非ゼロ終了する")
    args = ap.parse_args()

    apath = args.dir / "assignment.csv"
    if not apath.exists():
        sys.exit(f"[ERROR] {apath} がありません。先に make_screening_sheets.py を実行すること")
    with apath.open(encoding="utf-8-sig", newline="") as f:
        assignment = {r["record_id"]: r for r in csv.DictReader(f)}

    sheets = {rid: load_sheet(args.dir / f"sheet_{rid}.csv") for rid in REVIEWERS}
    missing_files = [r for r, s in sheets.items() if not s]
    if missing_files:
        print(f"[WARN] 判定シートが見つからない/空: {', '.join(missing_files)}")

    # --- 1. 検査 -----------------------------------------------------------
    problems: list[str] = []
    blank = Counter()
    for rev, sheet in sheets.items():
        # stage 1 の担当: 著者は全件、他2名は校正セットのみ
        expected = (set(assignment) if rev == "author"
                    else {rid for rid, a in assignment.items()
                          if a.get("calibration") == "Y"})
        got = set(sheet)
        if got - expected:
            problems.append(f"{rev}: 担当外の record_id が {len(got - expected)} 件")
        if expected - got:
            problems.append(f"{rev}: シートに無い担当レコードが {len(expected - got)} 件")
        for rid, row in sheet.items():
            d = (row.get("decision") or "").strip().lower()
            if not d:
                blank[rev] += 1
            elif d not in VALID:
                problems.append(f"{rev}/{rid}: 不正な decision '{row.get('decision')}'")
            elif d == "exclude" and not (row.get("reason") or "").strip():
                problems.append(f"{rev}/{rid}: Exclude だが reason が空")

    print("=" * 66)
    print("  Phase 3b 集計")
    print("=" * 66)
    print(f"  対象レコード: {len(assignment):,} 件")
    for rev in REVIEWERS:
        n_exp = (len(assignment) if rev == "author"
                 else sum(1 for a in assignment.values()
                          if a.get("calibration") == "Y"))
        # 「シートに行があって decision が妥当」なものだけを記入済みと数える。
        # 行そのものが欠けているケースを未記入に含めないと進捗が過大に出る。
        done = sum(1 for rid, row in sheets.get(rev, {}).items()
                   if rid in assignment
                   and (row.get("decision") or "").strip().lower() in VALID)
        pct = (done / n_exp * 100) if n_exp else 0
        print(f"    {REVIEWERS[rev]:18s}: {done:5,d}/{n_exp:5,d} 記入済 ({pct:5.1f}%)")

    if problems:
        print(f"\n  [問題 {len(problems)} 件]")
        for p in problems[:15]:
            print(f"    - {p}")
        if len(problems) > 15:
            print(f"    ... 他 {len(problems) - 15} 件")

    # --- 2. ペアごとの Cohen's κ(校正セットのみ) --------------------------
    #
    # ★ κ は **校正セット(calibration=Y、3名全員が全判定を行う)でのみ**算出する。
    #   liberal accelerated の除外プールで κ を計算してはならない。著者側の判定が
    #   定義上すべて Exclude で分散が無いため Pe = Po となり、
    #   **実際の一致率によらず κ が常に 0** になるためである(実測で確認済み)。
    pair_rows: dict[tuple[str, str], list[tuple[str, str]]] = defaultdict(list)
    cal_ids = [rid for rid, a in assignment.items() if a.get("calibration") == "Y"]
    for rid in cal_ids:
        decs = {}
        for rv in REVIEWERS:
            d = (sheets.get(rv, {}).get(rid, {}).get("decision") or "").strip().lower()
            if d in VALID:
                decs[rv] = CANON[d]
        for ra, rb in combinations(sorted(decs), 2):
            pair_rows[(ra, rb)].append((decs[ra], decs[rb]))

    print(f"\n  --- ペアごとの Cohen's κ(校正セット {len(cal_ids):,} 件・"
          f"3カテゴリ: Include/Exclude/Unsure) ---")
    kappas = []
    for pair in sorted(pair_rows):
        k, po, pe, n = cohen_kappa(pair_rows[pair])
        kappas.append(k)
        names = " × ".join(REVIEWERS[p] for p in pair)
        print(f"    {names:40s} n={n:5,d}  κ={k:6.3f}  "
              f"(Po={po:.3f} Pe={pe:.3f})  {landis_koch(k)}")
    valid_k = [k for k in kappas if k == k]
    if valid_k:
        mean_k = sum(valid_k) / len(valid_k)
        print(f"    {'平均':40s}          κ={mean_k:6.3f}  {landis_koch(mean_k)}"
              "   ← 本文で報告する値(Rev.9)")

    # 参考: Unsure を Include に寄せた2カテゴリ版
    print("\n  --- 参考: Unsure→Include とみなした2カテゴリ版 ---")
    k2s = []
    for pair in sorted(pair_rows):
        collapsed = [tuple("Include" if x == "Unsure" else x for x in p)
                     for p in pair_rows[pair]]
        k, _, _, n = cohen_kappa(collapsed)
        k2s.append(k)
        print(f"    {' × '.join(REVIEWERS[p] for p in pair):40s} n={n:5,d}  κ={k:6.3f}")
    valid_k2 = [k for k in k2s if k == k]
    if valid_k2:
        print(f"    {'平均':40s}          κ={sum(valid_k2)/len(valid_k2):6.3f}")

    # --- 3. 協議リスト -----------------------------------------------------
    worklist, agreed = [], []
    # stage 2(著者が Exclude/Unsure にした分の第2評価)。無ければ空。
    stage2 = {rv: load_sheet(args.dir / f"stage2_sheet_{rv}.csv")
              for rv in REVIEWERS if rv != "author"}
    n_liberal = 0
    for rid, a in assignment.items():
        ra = "author"
        row_a = sheets.get(ra, {}).get(rid, {})
        da = (row_a.get("decision") or "").strip().lower()

        if a.get("calibration") == "Y":
            # 校正セットは3名全員が判定済み。著者 × 割当上の第2評価者で突き合わせる。
            rb = "kataoka" if second_reviewer_of(a["key"]) == "kataoka" else "watanabe"
            row_b = sheets.get(rb, {}).get(rid, {})
        else:
            rb = a["reviewer_b"]
            # ★ liberal accelerated: Include は1名の判断で通す(第2評価者に回さない)。
            #   誤 Include は Phase 4 の手間が増えるだけだが、誤 Exclude は
            #   全文を読む機会が永久に失われるため、除外側にだけ2名を要求する。
            if da == "include":
                n_liberal += 1
                agreed.append({
                    "record_id": rid, "block": a["block"], "title": a["title"],
                    "doi": a["doi"], "reviewer_a": ra, "decision_a": "Include",
                    "reason_a": row_a.get("reason", ""),
                    "reviewer_b": "—", "decision_b": "—", "reason_b": "",
                    "final_decision": "Include",
                    "consensus_note": "liberal accelerated: 1名の Include で通過",
                })
                continue
            row_b = stage2.get(rb, {}).get(rid, {})

        db = (row_b.get("decision") or "").strip().lower()
        if da not in VALID or db not in VALID:
            continue
        rec = {
            "record_id": rid, "block": a["block"], "title": a["title"], "doi": a["doi"],
            "reviewer_a": ra, "decision_a": CANON[da], "reason_a": row_a.get("reason", ""),
            "reviewer_b": rb, "decision_b": CANON[db], "reason_b": row_b.get("reason", ""),
        }
        if da == db and da != "unsure":
            rec["final_decision"] = CANON[da]
            rec["consensus_note"] = "両者一致(協議不要)"
            agreed.append(rec)
        else:
            rec["final_decision"] = ""      # ← 協議で記入
            rec["consensus_note"] = ""      # ← 協議で記入
            worklist.append(rec)

    print("\n  --- 協議の要否 ---")
    print(f"    両者一致(協議不要): {len(agreed):5,d} 件")
    print(f"    要協議(不一致/Unsure): {len(worklist):5,d} 件")
    if agreed or worklist:
        dist = Counter(r["final_decision"] for r in agreed)
        print(f"      一致の内訳: Include {dist.get('Include', 0):,} / "
              f"Exclude {dist.get('Exclude', 0):,}")

    cols = ["record_id", "block", "title", "doi",
            "reviewer_a", "decision_a", "reason_a",
            "reviewer_b", "decision_b", "reason_b",
            "final_decision", "consensus_note"]

    args.dir.mkdir(parents=True, exist_ok=True)
    wpath = args.dir / "consensus_worklist.csv"
    if worklist:
        # 既存の協議結果を消さないよう、記入済みがあればマージする
        prev = load_sheet(wpath)
        for r in worklist:
            p = prev.get(r["record_id"], {})
            r["final_decision"] = (p.get("final_decision") or "").strip()
            r["consensus_note"] = (p.get("consensus_note") or "").strip()
        with wpath.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(worklist)
        print(f"\n  出力: {wpath}  ({len(worklist):,} 件)")
        print("        final_decision に協議結果を記入すること。")
        print("        **協議で解決しない場合は Include に倒す**(再現率優先、rule.md Phase 3b)。")

    # --- 4. 最終判定(協議が埋まっていれば) --------------------------------
    unresolved = [r for r in worklist
                  if (r.get("final_decision") or "").strip().lower() not in VALID]
    if worklist and unresolved:
        print(f"\n  [INFO] 未解決の協議 {len(unresolved):,} 件があるため "
              f"final_decisions.csv は生成しない。")
    else:
        final = agreed + worklist
        for r in final:
            r["final_decision"] = CANON[r["final_decision"].strip().lower()]
        fpath = args.dir / "final_decisions.csv"
        with fpath.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            w.writerows(final)
        dist = Counter(r["final_decision"] for r in final)
        print(f"\n  出力: {fpath}  ({len(final):,} 件)")
        print(f"        Include {dist.get('Include', 0):,} → Phase 4 へ / "
              f"Exclude {dist.get('Exclude', 0):,}")

    if args.strict and (problems or any(blank.values())):
        sys.exit(1)


if __name__ == "__main__":
    main()
