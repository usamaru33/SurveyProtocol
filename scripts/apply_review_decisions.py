#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""review_<id>.html が書き出した decisions CSV を判定シート(xlsx/csv)へ反映する。

検証してから書く。落ちる条件:
  - 担当外の record_id が混じっている
  - decision が Include/Exclude/Unsure 以外
  - Exclude なのに reason が空、または統制語彙の外
  - reason が「その他」なのに note が空
  - 既に別の判定が入っているセルを上書きしようとしている(--force で許可)

使い方:
    python -X utf8 scripts/apply_review_decisions.py --id author --input decisions_author.csv
    python -X utf8 scripts/apply_review_decisions.py --id author --input ... --dry-run
"""
from __future__ import annotations

import argparse
import csv
import shutil
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
csv.field_size_limit(10 ** 9)

from make_screening_xlsx import DECISIONS, REASON_OTHER, REASON_VALUES  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SCREENING = ROOT / "screening"
CANON = {d.lower(): d for d in DECISIONS}


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--id", default="author")
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--dry-run", action="store_true", help="検証だけして書かない")
    ap.add_argument("--force", action="store_true",
                    help="既存の判定を上書きする（既定は拒否）")
    args = ap.parse_args()

    xlsx = SCREENING / f"sheet_{args.id}.xlsx"
    if not xlsx.exists():
        sys.exit(f"[ERROR] {xlsx} が無い")
    if not args.input.exists():
        sys.exit(f"[ERROR] {args.input} が無い")

    with args.input.open(encoding="utf-8-sig", newline="") as f:
        incoming = {r["record_id"]: r for r in csv.DictReader(f) if r.get("record_id")}
    print(f"[INFO] 入力 {len(incoming):,} 件")

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("[ERROR] openpyxl が必要:  pip install openpyxl")

    wb = load_workbook(xlsx)
    ws = wb["判定"]
    header = [c.value for c in ws[1]]
    col = {n: header.index(n) + 1 for n in ("ID", "判定 ★", "除外理由 ★", "メモ")}

    errors: list[str] = []
    writes: list[tuple[int, str, str, str]] = []
    seen: set[str] = set()

    for row in range(2, ws.max_row + 1):
        rid = ws.cell(row, col["ID"]).value
        if not rid:
            continue
        rid = str(rid)
        got = incoming.get(rid)
        if not got:
            continue
        seen.add(rid)

        dec = CANON.get((got.get("decision") or "").strip().lower(), "")
        reason = (got.get("reason") or "").strip()
        note = (got.get("note") or "").strip()

        if not dec:
            errors.append(f"{rid}: decision が不正 ({got.get('decision')!r})")
            continue
        if dec == "Exclude":
            if not reason:
                errors.append(f"{rid}: Exclude なのに除外理由が空")
            elif reason not in REASON_VALUES:
                errors.append(f"{rid}: 除外理由が統制語彙の外 ({reason!r})")
            elif reason == REASON_OTHER and not note:
                errors.append(f"{rid}: 除外理由が「{REASON_OTHER}」なのにメモが空")
        else:
            reason = ""

        cur = (ws.cell(row, col["判定 ★"]).value or "")
        cur = str(cur).strip()
        if cur and cur != dec and not args.force:
            errors.append(f"{rid}: 既存の判定 {cur!r} を {dec!r} で上書きしようとしている"
                          f"（意図的なら --force）")
            continue
        writes.append((row, dec, reason, note))

    unknown = set(incoming) - seen
    if unknown:
        errors.append(f"担当外の record_id が {len(unknown)} 件: "
                      + ", ".join(sorted(unknown)[:5]) + (" …" if len(unknown) > 5 else ""))

    if errors:
        print(f"\n[NG] 検証エラー {len(errors)} 件:")
        for e in errors[:40]:
            print("   -", e)
        if len(errors) > 40:
            print(f"   … 他 {len(errors) - 40} 件")
        sys.exit(1)

    print(f"[OK] 検証を通過。書き込み対象 {len(writes):,} 件")
    if args.dry_run:
        print("[INFO] --dry-run のため書き込まない")
        return

    backup = xlsx.with_suffix(f".{datetime.now():%Y%m%d_%H%M%S}.bak.xlsx")
    shutil.copy2(xlsx, backup)
    print(f"[INFO] バックアップ: {backup.name}")

    for row, dec, reason, note in writes:
        ws.cell(row, col["判定 ★"]).value = dec
        ws.cell(row, col["除外理由 ★"]).value = reason or None
        if note:
            ws.cell(row, col["メモ"]).value = note
    wb.save(xlsx)

    filled = sum(1 for r in range(2, ws.max_row + 1)
                 if str(ws.cell(r, col["判定 ★"]).value or "").strip())
    total = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, col["ID"]).value)
    print(f"[OK] {xlsx.name} に反映した。記入済み {filled:,}/{total:,}")
    print("     集計は  python -X utf8 scripts/score_screening.py")


if __name__ == "__main__":
    main()
